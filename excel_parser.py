"""Excel-LV-Parser für Ausschreibungen ohne feste Struktur.

Erkennt Kopfzeile, Spaltenrollen und Zeilentypen (Gruppe / Position / Hinweis) und
liefert ein ``GaebProject`` — damit greift die komplette GAEB-Nachverarbeitung
(Matching, ``_import_gaeb_groups``, Entwürfe, Easyjob-Anlage) unverändert.

Die Erkennung ist nur ein Vorschlag: ``probe_workbook`` liefert zusätzlich die
Vorschauzeilen für den Mapping-Dialog, in dem der Nutzer per Farb-Pinsel korrigiert.
Das korrigierte ``ExcelLayout`` geht dann an ``parse_excel``.
"""
import hashlib
import io
import re
import threading
from collections import OrderedDict
from dataclasses import dataclass, field, asdict
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from gaeb_parser import GaebItem, GaebProject, GaebRemark

# ─── Rollen (Spalten) und Zeilentypen ────────────────────────────────────────

ROLE_IGNORE = "ignore"
ROLE_OZ     = "oz"       # Positionsnummer / Ordnungszahl
ROLE_DESC   = "desc"     # Beschreibung / Leistungsbeschreibung
ROLE_REF    = "ref"      # Referenz / Typ / Hersteller / Bemerkung
ROLE_UNIT   = "unit"     # Einheit
ROLE_QTY    = "qty"      # Menge / Anzahl / Stückzahl
ROLE_PRICE  = "price"    # Einzelpreis (Ziel der Rückschreibung)
ROLE_TOTAL  = "total"    # Gesamtpreis (Menge x Einzelpreis) — nur wenn zugeordnet
ROLE_FLAG   = "flag"     # AP/BP — Alternativ-/Bedarfsposition

COLUMN_ROLES = (ROLE_OZ, ROLE_DESC, ROLE_REF, ROLE_UNIT, ROLE_QTY, ROLE_PRICE,
                ROLE_TOTAL, ROLE_FLAG, ROLE_IGNORE)

ROW_HEADER = "header"
ROW_JOB    = "job"       # eigener Easyjob-Job
ROW_MAIN   = "main"      # Hauptgruppe (Ebene 1 im Job)
ROW_GROUP  = "grp"       # Gruppe (Ebene 2 und tiefer)
ROW_POS    = "pos"
ROW_NOTE   = "note"
ROW_SKIP   = "skip"

ROW_KINDS = (ROW_HEADER, ROW_JOB, ROW_MAIN, ROW_GROUP, ROW_POS, ROW_NOTE, ROW_SKIP)
ROW_GROUPISH = (ROW_JOB, ROW_MAIN, ROW_GROUP)

# Ebenen der gemalten Gruppen-Pinsel — bewusst fest, damit „Hauptgruppe" immer
# dasselbe bedeutet und nicht von der Nachbarzeile abhängt. ROW_GROUP wird nicht
# mehr angeboten (in Easyjob ist eine Gruppe dasselbe wie eine Position), bleibt
# hier aber gültig, damit gespeicherte Layouts weiter funktionieren.
_PAINT_LEVEL = {ROW_JOB: 0, ROW_MAIN: 1, ROW_GROUP: 2}

# Blätter werden zu … (nur relevant, wenn mehrere Blätter aktiv sind)
SHEET_AS_JOB  = "job"    # jedes Blatt ein eigener Easyjob-Job
SHEET_AS_KEEP = "keep"   # ein Job für alles, Hauptgruppen des Blatts bleiben erhalten
SHEET_AS_MAIN = "main"   # ein Job, das Blatt IST die Hauptgruppe — Gruppen darin
                         # werden Hinweise (Easyjob kennt keine Ebene darunter)
SHEET_MODES   = (SHEET_AS_JOB, SHEET_AS_KEEP, SHEET_AS_MAIN)

# Synonyme für die Kopfzeilen-Erkennung (Präfix-Vergleich auf normalisiertem Text)
_HEADER_WORDS: dict[str, tuple[str, ...]] = {
    ROLE_OZ:    ("pos", "nr", "positionsnummer", "positions-nr", "ordnungszahl", "oz", "lfd"),
    ROLE_DESC:  ("beschreibung", "bezeichnung", "leistungsbeschreibung", "leistung",
                 "artikel", "text", "description"),
    ROLE_REF:   ("referenz", "bemerkung", "bemerkungen", "typ", "gerätetyp", "geraetetyp",
                 "hersteller", "kommentar", "fabrikat", "reference"),
    ROLE_UNIT:  ("einheit", "einh", "me", "mengeneinheit", "unit"),
    ROLE_QTY:   ("menge", "anzahl", "anz", "stückzahl", "stueckzahl", "stk", "stck",
                 "qty", "quantity"),
    ROLE_PRICE: ("einzelpreis", "ep", "preis", "einheitspreis", "unit price"),
    ROLE_TOTAL: ("gesamtpreis", "gesamtbetrag", "gesamt", "total", "summe",
                 "total netto", "total price"),
    ROLE_FLAG:  ("ap/bp", "positionsart", "art der position", "positionstyp"),
}

# Zeilen, die nur Summen aufaddieren — keine Positionen, keine Gruppen
_SUM_RE = re.compile(
    r'^(zwischen|teil|gesamt|end)?summe\b|^summe$|^total\b|^gesamt(preis|betrag)?$'
    r'|^übertrag|^uebertrag', re.IGNORECASE)

# Positionsnummer: "1.", "1.1", "01.01.01.01", "A.01", "B-12", "2.1.14"
_POSNO_RE = re.compile(r'^[A-Za-z]{0,2}[.\-]?\d+([.\-]\d+)*\.?$')

# Führende Aufzählungsstriche = Unterüberschrift (LOS1 GOE: "-- UHD Professional …")
_DASH_RE = re.compile(r'^([-–—•·]+)\s*')

_ALT_RE  = re.compile(r'^(ap|a)$|^al?ternativ', re.IGNORECASE)
_EVENT_RE = re.compile(r'^(bp|b)$|^bedarf|^eventual', re.IGNORECASE)

_NL = chr(10)               # Zeilentrenner in Langtexten
_MAX_HEADER_SCAN = 25    # Zeilen, in denen die Kopfzeile gesucht wird
_MAX_COL_SCAN    = 60    # Spalten-Obergrenze (PAG 01_Material meldet max_column=16384)
_PREVIEW_ROWS    = 60      # Datenzeilen im Mapping-Dialog (Standard)
_ALL_ROWS        = 100000  # „alle Zeilen anzeigen" — praktisch unbegrenzt


def _rows_cap(sheet: str, show_all) -> int:
    return _ALL_ROWS if sheet in (show_all or ()) else _PREVIEW_ROWS


def _preview_cap(sheet: str, show_all, opened) -> int:
    """Zeilen, die für dieses Blatt in die Vorschau gehen.

    ``opened=None`` heißt „alle Blätter aufbauen" (erster Aufruf, da weiß der Server
    noch nicht, was der Browser offen hat). Sonst bekommen nur die aufgeklappten
    Blätter Zeilen: bei LOS2 gehören 92 % der Tabellenzellen zu zugeklappten Blättern,
    die der Nutzer nicht sieht und die der Browser bei jedem Pinselstrich neu aufbaut.
    """
    if opened is not None and sheet not in opened:
        return 0
    return _rows_cap(sheet, show_all)


# ─── Layout-Datenmodell (komplett JSON-serialisierbar) ───────────────────────

@dataclass
class QtyCol:
    """Eine Mengenspalte. Mehrere aktive Spalten = mehrere Szenarien = mehrere Jobs."""
    col:      int
    label:    str = ""       # aus den Zeilen über der Kopfzeile, z.B. "Event 3 · L · Valencia"
    job_name: str = ""       # Jobname in Easyjob (vorbelegt aus label)
    active:   bool = True
    values:   int = 0        # Anzahl Zahlenwerte im Datenbereich (nur Anzeige)


@dataclass
class ColumnRoles:
    """1-basierte Spaltenindizes; 0 = nicht gesetzt."""
    oz:    int = 0
    desc:  int = 0
    unit:  int = 0
    price: int = 0
    total: int = 0
    flag:  int = 0
    ref:   list[int]    = field(default_factory=list)
    qty:   list[QtyCol] = field(default_factory=list)

    def role_of(self, col: int) -> str:
        if col == self.desc:  return ROLE_DESC
        if col == self.oz:    return ROLE_OZ
        if col == self.unit:  return ROLE_UNIT
        if col == self.price: return ROLE_PRICE
        if col == self.total: return ROLE_TOTAL
        if col == self.flag:  return ROLE_FLAG
        if col in self.ref:   return ROLE_REF
        if any(q.col == col for q in self.qty): return ROLE_QTY
        return ROLE_IGNORE

    def active_qty(self) -> list[QtyCol]:
        return [q for q in self.qty if q.active]


@dataclass
class SheetLayout:
    name:           str
    sheet_no:       int                 # 1-basierte Position im Workbook
    enabled:        bool = True
    header_row:     int  = 0            # 0 = keine Kopfzeile erkannt
    roles:          ColumnRoles = field(default_factory=ColumnRoles)
    row_overrides:  dict[str, str] = field(default_factory=dict)  # "18" → ROW_*
    cell_overrides: dict[str, str] = field(default_factory=dict)  # "18:4" → ROLE_*

    @property
    def first_data_row(self) -> int:
        return self.header_row + 1 if self.header_row else 1


@dataclass
class ExcelLayout:
    sheets:      list[SheetLayout] = field(default_factory=list)
    fingerprint: str = ""
    label:       str = ""
    # Was ein Blatt wird, wenn mehrere aktiv sind — begrenzt die Anzahl der Easyjob-Jobs.
    sheet_mode:  str = SHEET_AS_JOB

    def sheet(self, name: str) -> Optional[SheetLayout]:
        return next((s for s in self.sheets if s.name == name), None)


@dataclass
class PreviewRow:
    row:   int
    kind:  str
    level: int
    cells: list[str]        # Index 0 = Spalte 1


@dataclass
class SheetProbe:
    layout:  SheetLayout
    max_col: int
    preview: list[PreviewRow]
    counts:  dict[str, int]     # {"pos": n, "grp": n, "note": n}
    total_rows: int
    header_context: list[list[str]] = field(default_factory=list)  # bis zu 3 Zeilen über der Kopfzeile
    header_texts:   list[str]       = field(default_factory=list)  # Kopfzeile selbst (Fingerprint)


@dataclass
class WorkbookProbe:
    sheets:       list[SheetProbe]
    fingerprint:  str
    project_name: str

    sheet_mode:   str = SHEET_AS_JOB

    @property
    def layout(self) -> ExcelLayout:
        return ExcelLayout(sheets=[sp.layout for sp in self.sheets],
                           fingerprint=self.fingerprint, sheet_mode=self.sheet_mode)

    @property
    def active_sheets(self) -> int:
        return sum(1 for sp in self.sheets if sp.layout.enabled)


# ─── Serialisierung (Entwürfe + Layout-Profile in der DB) ────────────────────

def layout_to_dict(layout: ExcelLayout) -> dict:
    return asdict(layout)


def layout_from_dict(data: dict) -> ExcelLayout:
    sheets = []
    for s in data.get("sheets") or []:
        r = s.get("roles") or {}
        roles = ColumnRoles(
            oz=int(r.get("oz") or 0), desc=int(r.get("desc") or 0),
            unit=int(r.get("unit") or 0), price=int(r.get("price") or 0),
            total=int(r.get("total") or 0), flag=int(r.get("flag") or 0),
            ref=[int(c) for c in (r.get("ref") or [])],
            qty=[QtyCol(col=int(q.get("col") or 0), label=q.get("label") or "",
                        job_name=q.get("job_name") or "",
                        active=bool(q.get("active", True)), values=int(q.get("values") or 0))
                 for q in (r.get("qty") or []) if q.get("col")],
        )
        sheets.append(SheetLayout(
            name=s.get("name") or "", sheet_no=int(s.get("sheet_no") or 0),
            enabled=bool(s.get("enabled", True)), header_row=int(s.get("header_row") or 0),
            roles=roles,
            row_overrides={str(k): v for k, v in (s.get("row_overrides") or {}).items()},
            cell_overrides={str(k): v for k, v in (s.get("cell_overrides") or {}).items()},
        ))
    mode = data.get("sheet_mode") or SHEET_AS_JOB
    return ExcelLayout(sheets=sheets, fingerprint=data.get("fingerprint") or "",
                       label=data.get("label") or "",
                       sheet_mode=mode if mode in SHEET_MODES else SHEET_AS_JOB)


# ─── Zell-Helfer ─────────────────────────────────────────────────────────────

def _norm(v: Any) -> str:
    """Normalisiert Zelltext für Vergleiche: klein, Zeilenumbrüche/Mehrfach-Spaces weg."""
    if v is None:
        return ""
    return re.sub(r'\s+', ' ', str(v).strip().lower())


def _txt(v: Any) -> str:
    """Zelltext für die Ausgabe: Zeilenumbrüche bleiben, Rand-Whitespace weg."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _num(v: Any) -> Optional[float]:
    """Zahl aus einer Zelle — auch wenn sie als Text formatiert ist.

    Trennzeichen sind mehrdeutig, deshalb eine feste Regel:
      * Kommt ein **Komma** vor, ist es das Dezimalzeichen und Punkte sind
        Tausenderpunkte:            "1.234,50" -> 1234.5
      * Sonst ist ein **einzelner Punkt mit genau drei Ziffern dahinter** ein
        Tausenderpunkt (deutsche Schreibweise):  "1.234" -> 1234.0
      * Jeder andere Punkt ist ein Dezimalpunkt: "1.5" -> 1.5, "12.75" -> 12.75

    Ohne diese Unterscheidung fiel früher jede als Text formatierte Dezimalmenge auf
    ihren ganzzahligen Teil zurück ("0,5" -> 0), weil die Tausender-Alternative im
    Regex zuerst griff.
    """
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    m = re.search(r'-?\d+(?:[.,\s ]\d+)*', s)
    if not m:
        return None
    raw = m.group(0).replace(" ", "").replace(" ", "")
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        parts = raw.split(".")
        if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3):
            raw = raw.replace(".", "")      # Tausenderpunkte
    try:
        return float(raw)
    except ValueError:
        return None


def _fill_key(cell) -> tuple:
    """Vergleichbare Signatur der Füllfarbe. openpyxl liefert je nach Typ
    theme+tint, rgb oder indexed — .rgb blind lesen wirft/liefert Müll."""
    f = getattr(cell, "fill", None)
    if f is None or f.patternType != "solid":
        return ("none",)
    c = f.fgColor
    if c is None:
        return ("none",)
    if c.type == "theme":
        return ("theme", c.theme, round(float(c.tint or 0.0), 3))
    if c.type == "rgb":
        return ("rgb", c.rgb if isinstance(c.rgb, str) else "")
    if c.type == "indexed":
        return ("indexed", c.indexed)
    return ("other",)


def _fill_darkness(key: tuple) -> float:
    """0 = kein/hellster Hintergrund, 1 = dunkelster. Rangfolge für Gruppenebenen."""
    if key[0] == "theme":
        # Negativer Tint = dunkler. Theme 1/3 sind die Text-/Akzentfarben (dunkel).
        return -float(key[2]) + (0.5 if key[1] in (1, 3) else 0.0)
    if key[0] == "rgb":
        s = key[1] or ""
        if len(s) == 8:
            s = s[2:]
        if len(s) == 6:
            try:
                r, g, b = int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)
                return 1.0 - (0.299 * r + 0.587 * g + 0.114 * b) / 255.0
            except ValueError:
                return 0.5
        return 0.5
    if key[0] == "none":
        return 0.0
    return 0.5


def _style_sig(cell) -> tuple:
    """Signatur für die Gruppenebenen-Rangfolge: Füllung, fett, Schriftgröße, Einzug."""
    fnt = getattr(cell, "font", None)
    aln = getattr(cell, "alignment", None)
    return (
        _fill_key(cell),
        bool(fnt and fnt.bold),
        round(float(fnt.size or 11.0), 1) if fnt and fnt.size else 11.0,
        round(float(aln.indent or 0.0), 1) if aln else 0.0,
    )


def _is_styled(cell) -> bool:
    """Optisch als Überschrift hervorgehoben (fett oder farbig hinterlegt)."""
    fnt = getattr(cell, "font", None)
    return bool((fnt and fnt.bold) or _fill_key(cell)[0] != "none")


def _looks_like_posno(text: str) -> bool:
    t = text.strip()
    return bool(t) and bool(_POSNO_RE.match(t)) and any(ch.isdigit() for ch in t)


def _posno_depth(text: str) -> int:
    """Hierarchietiefe einer Positionsnummer: '01' → 1, '01.01' → 2, 'A.01' → 1."""
    t = text.strip().rstrip(".")
    parts = [p for p in re.split(r'[.\-]', t) if p]
    # Ein reiner Buchstaben-Präfix ("A" in "A.01") ist Bereichskennung, keine Ebene
    if parts and parts[0].isalpha():
        parts = parts[1:]
    return max(1, len(parts))


def _used_max_col(ws, upto_row: int) -> int:
    """Letzte tatsächlich befüllte Spalte — ws.max_column ist bei durchformatierten
    Sheets 16384 (PAG 01_Material)."""
    limit = min(ws.max_column or 1, _MAX_COL_SCAN)
    last = 1
    for r in range(1, min(ws.max_row or 1, upto_row) + 1):
        for c in range(limit, last, -1):
            if ws.cell(r, c).value is not None:
                last = c
                break
    return last


# ─── Kopfzeilen- und Spalten-Erkennung ──────────────────────────────────────

def _match_role(text: str) -> Optional[str]:
    """Ordnet einen Kopfzeilentext einer Rolle zu (längstes passendes Synonym gewinnt)."""
    n = _norm(text)
    if not n:
        return None
    best: tuple[int, str] | None = None
    for role, words in _HEADER_WORDS.items():
        for w in words:
            if n == w or n.startswith(w + " ") or n.startswith(w + ".") or n.startswith(w + "\n"):
                if best is None or len(w) > best[0]:
                    best = (len(w), role)
    return best[1] if best else None


def _detect_header_row(ws, max_col: int) -> tuple[int, dict[str, list[int]]]:
    """Kopfzeile = Zeile mit den meisten Rollen-Treffern (desc doppelt gewichtet)."""
    best: tuple[int, int, dict[str, list[int]]] = (-1, 0, {})
    for r in range(1, min(ws.max_row or 1, _MAX_HEADER_SCAN) + 1):
        hits: dict[str, list[int]] = {}
        for c in range(1, max_col + 1):
            role = _match_role(ws.cell(r, c).value)
            if role:
                hits.setdefault(role, []).append(c)
        if not hits:
            continue
        score = len(hits) + (2 if ROLE_DESC in hits else 0) + (1 if ROLE_QTY in hits else 0)
        if score > best[0]:
            best = (score, r, hits)
    return (best[1], best[2]) if best[0] > 0 else (0, {})


def _qty_label(ws, header_row: int, col: int) -> str:
    """Szenario-Label aus den bis zu 3 Zeilen über der Kopfzeile — bei mehreren
    Mengenspalten steht dort, wofür sie stehen (GOE: 'Event 3' / 'L' / 'Valencia')."""
    parts: list[str] = []
    for r in range(max(1, header_row - 4), header_row):
        # Mengen- und zugehörige Preisspalte liegen nebeneinander; die Überschrift
        # steht mal über der einen, mal über der anderen ("Event 1" | "S").
        words = []
        for c in (col, col + 1):
            v = re.sub(r'\s+', ' ', _txt(ws.cell(r, c).value))
            # Reine Zahlen sind Tagesangaben (Auf-/Abbau, VA-Tage), keine Bezeichnung
            if not v or re.fullmatch(r'-?[\d.,\s]+', v):
                continue
            words.append(v)
        line = " ".join(words)[:40].strip()
        if line and line not in parts:
            parts.append(line)
    return " · ".join(parts[:3])


def _detect_ref_columns(ws, header_row: int, max_col: int,
                        roles: ColumnRoles, hits: dict[str, list[int]]) -> list[int]:
    """Referenz-/Typspalten. Neben den Kopfzeilen-Treffern auch die header-lose
    Spalte direkt neben der Beschreibung (Vector: Spalte D mit 'FD34, H30V')."""
    refs = list(hits.get(ROLE_REF, []))
    taken = {roles.desc, roles.oz, roles.unit, roles.price, roles.total, roles.flag}
    taken |= {q.col for q in roles.qty}
    refs = [c for c in refs if c not in taken]

    for cand in (roles.desc + 1, roles.desc - 1):
        if cand < 1 or cand > max_col or cand in taken or cand in refs:
            continue
        if _txt(ws.cell(header_row, cand).value):
            continue                       # hat einen Header → wäre oben erkannt worden
        texts, numbers = 0, 0
        for r in range(header_row + 1, min(ws.max_row or 1, header_row + 200) + 1):
            v = ws.cell(r, cand).value
            if v is None:
                continue
            if isinstance(v, (int, float)):
                numbers += 1
            elif len(str(v).strip()) <= 40:
                texts += 1
        if texts >= 3 and texts > numbers:
            refs.append(cand)
    return sorted(refs)


def detect_sheet_layout(ws, sheet_no: int) -> tuple[SheetLayout, int]:
    """Auto-Erkennung für ein Sheet. Gibt (Layout, genutzte Spaltenzahl) zurück."""
    max_col = _used_max_col(ws, _MAX_HEADER_SCAN + _PREVIEW_ROWS)
    header_row, hits = _detect_header_row(ws, max_col)
    roles = ColumnRoles()
    if not header_row:
        return SheetLayout(name=ws.title, sheet_no=sheet_no, enabled=False,
                           header_row=0, roles=roles), max_col

    roles.desc  = (hits.get(ROLE_DESC)  or [0])[0]
    roles.oz    = (hits.get(ROLE_OZ)    or [0])[0]
    roles.unit  = (hits.get(ROLE_UNIT)  or [0])[0]
    roles.price = (hits.get(ROLE_PRICE) or [0])[0]
    roles.total = (hits.get(ROLE_TOTAL) or [0])[0]
    roles.flag  = (hits.get(ROLE_FLAG)  or [0])[0]

    # Alle Mengenspalten, nicht nur die erste (GOE: H/J/L/N je Event)
    qty_cols = hits.get(ROLE_QTY) or []
    for c in qty_cols:
        values = sum(1 for r in range(header_row + 1, min(ws.max_row or 1, header_row + 400) + 1)
                     if isinstance(ws.cell(r, c).value, (int, float)))
        lbl = _qty_label(ws, header_row, c) if len(qty_cols) > 1 else ""
        roles.qty.append(QtyCol(col=c, label=lbl, job_name=lbl, values=values))

    roles.ref = _detect_ref_columns(ws, header_row, max_col, roles, hits)
    enabled = bool(roles.desc or roles.oz)
    return SheetLayout(name=ws.title, sheet_no=sheet_no, enabled=enabled,
                       header_row=header_row, roles=roles), max_col


# ─── Zeilen-Klassifikation ──────────────────────────────────────────────────

@dataclass
class _Row:
    """Eine ausgewertete Datenzeile."""
    row:      int
    kind:     str
    level:    int          # nur bei ROW_GROUP relevant
    label:    str          # Gruppenname bzw. Positions-Kurztext (erste Zeile)
    long:     str          # restliche Zeilen der Beschreibung
    oz:       str
    unit:     str
    ref:      str
    flag:     str
    qty:      dict[int, Optional[float]]   # Spalte → Menge
    style:    tuple
    painted:  bool = False   # Ebene kommt aus dem Pinsel, nicht aus der Erkennung
    has_pos:  bool = False   # Gruppe enthält (äußerste Ebene) Positionen → Hauptgruppe
    was_group: bool = False  # als Hinweis herabgestuft, gehört aber in die Gruppenkette


def _cell_for(layout: SheetLayout, row: int, role: str, default_col: int) -> int:
    """Einzelzellen-Override berücksichtigen: '18:4' → ROLE_DESC heißt, in Zeile 18
    kommt die Beschreibung aus Spalte 4."""
    for key, r in layout.cell_overrides.items():
        if r != role:
            continue
        try:
            kr, kc = key.split(":")
        except ValueError:
            continue
        if int(kr) == row:
            return int(kc)
    return default_col


def _read_row(ws, layout: SheetLayout, row: int) -> _Row:
    roles = layout.roles
    d_col = _cell_for(layout, row, ROLE_DESC, roles.desc)
    o_col = _cell_for(layout, row, ROLE_OZ,   roles.oz)
    u_col = _cell_for(layout, row, ROLE_UNIT, roles.unit)
    f_col = _cell_for(layout, row, ROLE_FLAG, roles.flag)

    desc_raw = _txt(ws.cell(row, d_col).value) if d_col else ""
    oz_raw   = _txt(ws.cell(row, o_col).value) if o_col else ""
    unit     = re.sub(r'\s+', ' ', _txt(ws.cell(row, u_col).value)) if u_col else ""
    flag     = _txt(ws.cell(row, f_col).value) if f_col else ""
    ref      = " · ".join(t for t in (_txt(ws.cell(row, c).value) for c in roles.ref) if t)

    qty: dict[int, Optional[float]] = {}
    for q in roles.qty:
        qty[q.col] = _num(ws.cell(row, q.col).value)

    # Gruppenlabel kann in der Positionsnummer-Spalte stehen (LOS2: 'Licht' in Spalte B)
    from_oz = False
    label_src = desc_raw
    if not label_src and oz_raw and not _looks_like_posno(oz_raw):
        label_src, from_oz = oz_raw, True
    oz = oz_raw if (oz_raw and _looks_like_posno(oz_raw)) else ""

    lines  = [ln.strip() for ln in label_src.splitlines() if ln.strip()]
    label  = lines[0] if lines else ""
    long   = "\n".join(lines[1:])

    dash = 0
    m = _DASH_RE.match(label)
    if m:
        dash = len(m.group(1))
        label = label[m.end():].strip()

    style_cell = ws.cell(row, o_col if from_oz else (d_col or 1))
    style = _style_sig(style_cell)

    # ── Klassifikation ──
    kind, level, painted = ROW_NOTE, 0, False
    has_unit = bool(unit)
    has_qty  = any(v is not None for c, v in qty.items()
                   if any(q.col == c and q.active for q in roles.qty))
    styled   = _is_styled(style_cell)

    if not label:
        kind = ROW_SKIP
    elif _SUM_RE.match(label):
        kind = ROW_SKIP
    elif dash:
        kind, level = ROW_GROUP, -dash        # negativ = relativ zur Elterngruppe
    elif has_unit:
        kind = ROW_POS
    elif from_oz:
        kind = ROW_GROUP
    elif styled:
        kind = ROW_GROUP
    elif has_qty:
        kind = ROW_POS
    elif oz:
        kind = ROW_POS
    if kind == ROW_GROUP and level == 0 and oz:
        level = _posno_depth(oz)

    ov = layout.row_overrides.get(str(row))
    if ov in ROW_KINDS:
        kind = ov
        # Gemalte Ebenen sind absolut: „Hauptgruppe" heißt immer Ebene 1, egal was
        # darüber steht. Sonst hinge das Ergebnis von der Nachbarzeile ab.
        if kind in _PAINT_LEVEL:
            level   = _PAINT_LEVEL[kind]
            painted = True

    return _Row(row=row, kind=kind, level=level, label=label, long=long, oz=oz,
                unit=unit, ref=ref, flag=flag, qty=qty, style=style, painted=painted)


def _classify_sheet(ws, layout: SheetLayout, demote_main: bool = False) -> list[_Row]:
    """Alle Datenzeilen klassifizieren und die Gruppenebenen auflösen.

    demote_main: Das Blatt selbst ist die Hauptgruppe (sheet_mode == main). Dann kann
    keine Zeile darin ebenfalls Hauptgruppe sein — sie rutscht eine Ebene tiefer.
    """
    rows = [_read_row(ws, layout, r)
            for r in range(layout.first_data_row, (ws.max_row or 0) + 1)]
    rows = [r for r in rows if r.kind != ROW_SKIP]

    # Gemalte Jobs/Hauptgruppen haben ihre Ebene schon fest — nur erkannte Gruppen
    # (kind == ROW_GROUP mit level 0 bzw. negativ) müssen noch aufgelöst werden.
    # Strich-Tiefe normalisieren: nutzt eine Datei durchgehend "--", ist das *eine*
    # Ebene und nicht zwei (LOS1 GOE). Erst bei gemischtem "-"/"--" zählt der Unterschied.
    dashes = [-r.level for r in rows
              if r.kind == ROW_GROUP and r.level < 0 and not r.painted]
    dash_unit = min(dashes) if dashes else 1

    # Gruppen ohne Positionsnummer und ohne Strich: Stil-Rangfolge nach Dunkelheit
    unranked = {r.style for r in rows
                if r.kind == ROW_GROUP and r.level == 0 and not r.painted}
    rank: dict[tuple, int] = {}
    if unranked:
        sigs = sorted(unranked, key=lambda s: (-_fill_darkness(s[0]), -s[2], s[3]))
        rank = {s: i + 1 for i, s in enumerate(sigs)}

    # Strich- und Stilgruppen sind *relativ* zur umschließenden nummerierten Gruppe —
    # sonst landen "Traversen" (Stil) und "1. Rigging" (Nummer) beide auf Ebene 1.
    base = 0
    for r in rows:
        if r.kind not in ROW_GROUPISH:
            continue
        if r.painted:
            # Gemalte Ebenen sind absolut. Würden sie `base` verschieben, änderte ein
            # einzelner Pinselstrich die Ebene aller folgenden erkannten Gruppen.
            continue
        if r.level > 0:                              # Ebene aus der Positionsnummer
            base = r.level
        elif r.level < 0:                            # Strich-Untergruppe
            r.level = base + max(1, round(-r.level / dash_unit))
        else:                                        # Stil-Untergruppe
            r.level = base + rank.get(r.style, 1)

    # Erkannte Ebene 1 ist eine Hauptgruppe, alles darunter eine Gruppe — damit die
    # Vorschaufarben denselben Begriffen folgen wie die Pinsel.
    # Easyjob kennt genau eine Ebene zwischen Job und Position: die Hauptgruppe.
    # Also wird die *äußerste* Gruppe über einer Position deren Hauptgruppe — das sind
    # die großen Kategorien („Rigging", „Beleuchtung", „01 Videotechnik"), die auch in
    # Easyjob als Überschrift Sinn ergeben. Die feineren Ebenen darunter („Haupt-Rig",
    # „Traversen", „-- UHD Professional") werden Hinweise: sie bleiben an ihrer Stelle
    # sichtbar, erfinden aber keine Ebene, die es in Easyjob nicht gibt.
    stack: list[_Row] = []
    for r in rows:
        if r.kind in (ROW_JOB, ROW_MAIN, ROW_GROUP):
            if r.kind == ROW_JOB:
                stack = []          # neuer Job → neue äußerste Ebene
                continue
            while stack and stack[-1].level >= r.level:
                stack.pop()
            stack.append(r)
        elif r.kind == ROW_POS and stack:
            stack[0].has_pos = True

    for r in rows:
        if r.kind not in (ROW_MAIN, ROW_GROUP):
            continue
        if demote_main or not r.has_pos:
            # demote_main: das Blatt ist die Hauptgruppe, darunter gibt es keine Ebene.
            # was_group: der Text wird Hinweis, bleibt aber Teil der Gruppenkette für
            # das Matching — sonst verliert der Matcher „Traversen", „Displays" usw.
            r.kind      = ROW_NOTE
            r.was_group = True
        else:
            r.kind = ROW_MAIN
    return rows


def formula_rows(data: bytes) -> dict[str, dict[int, set[int]]]:
    """Zeilennummern der Formelzellen je Spalte: {Blatt: {Spalte: {Zeilen}}}.

    Wird gebraucht, um im Mapping-Dialog zu warnen, wenn die als Einzelpreis erkannte
    Spalte von der Datei selbst berechnet wird (LOS1 GOE: E = C + D). Dort darf nicht
    hineingeschrieben werden — besser jetzt sagen als erst beim Export. Zeilennummern
    statt bloßer Anzahl, weil nur Formeln in echten Positionszeilen zählen: Summen- und
    Zwischensummenzeilen sind normal und dürfen nicht warnen (LOS2 hat davon je Blatt
    ein paar in derselben Spalte).

    Eigener Ladevorgang, weil der Cache mit data_only=True nur die zwischengespeicherten
    Werte kennt, nicht die Formeln. read_only hält den Speicher klein (2–15 MB statt
    5–45 MB); das Ergebnis wird pro Upload einmal berechnet, nicht bei jedem Pinselstrich.
    """
    out: dict[str, dict[int, set[int]]] = {}
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    try:
        for ws in wb.worksheets:
            cols: dict[int, set[int]] = {}
            # Nur bis _MAX_COL_SCAN: weiter rechts kann das Layout keine Spalte
            # adressieren. PAGs 01_Material meldet max_column=16384 (eine
            # durchformatierte Zeile) — ungekappt sind das 3,9 Mio. Zellen statt 17.000.
            for r, row in enumerate(ws.iter_rows(values_only=True, max_col=_MAX_COL_SCAN),
                                    start=1):
                for idx, v in enumerate(row, start=1):
                    if isinstance(v, str) and v.startswith("="):
                        cols.setdefault(idx, set()).add(r)
            if cols:
                out[ws.title] = cols
    finally:
        wb.close()
    return out


# ─── Probe (Vorschau für den Mapping-Dialog) ────────────────────────────────

# Beim Zuordnen im Mapping-Dialog wird dieselbe Datei bei jedem Pinselstrich erneut
# gelesen. Das Laden dominiert die Antwortzeit (0,2–0,5 s je Mappe), deshalb wird die
# zuletzt benutzte Mappe gehalten. Bewusst klein: eine Mappe kostet je nach Datei
# 5–45 MB (PAG meldet max_column=16384). Nur lesende Nutzung — der Excel-Writer lädt
# seine eigene Kopie mit data_only=False.
_WB_CACHE: "OrderedDict[str, Any]" = OrderedDict()
_WB_CACHE_MAX  = 2
# Der Cache wird aus dem Event-Loop UND aus Executor-Threads angefasst. Ohne Lock
# kann ein Thread mitten in der Verdrängung stehen, während ein anderer einträgt →
# "RuntimeError: dictionary changed size during iteration". Die Schwestertabelle in
# routes/projects.py nimmt aus demselben Grund _OVERVIEW_CACHE_LOCK.
_WB_CACHE_LOCK = threading.Lock()
# Der Fingerprint hängt nur an den Bytes, wird aber bei jedem Pinselstrich neu
# gerechnet (3779 _match_role-Aufrufe je Request bei LOS2, 55 % der Repreview-Zeit)
# — und im Repreview überhaupt nicht gelesen. Also einmal je Datei merken.
_FP_CACHE: "OrderedDict[str, str]" = OrderedDict()
_FP_CACHE_MAX = 8      # nur Strings, darf großzügiger sein als der Mappen-Cache


def _wb_key(data: bytes) -> str:
    return hashlib.sha1(data).hexdigest()


def _load(data: bytes):
    """Arbeitsmappe lesen, zuletzt benutzte gecacht.

    Verdrängt wird die am längsten unbenutzte (LRU). Mit reinem FIFO flog beim
    Wechsel zwischen zwei Dateien genau die heraus, die gerade gebraucht wurde.

    Die Grenze ist bewusst niedrig: eine Mappe kostet je nach Datei 2–47 MB. Bei mehr
    als zwei gleichzeitigen Mapping-Sitzungen wird also wieder geladen — bekannt und
    in Kauf genommen, eine speicherbasierte Grenze wäre die nächste Stufe.
    """
    key = _wb_key(data)
    with _WB_CACHE_LOCK:
        wb = _WB_CACHE.get(key)
        if wb is not None:
            _WB_CACHE.move_to_end(key)      # frisch benutzt → nicht als nächstes raus
            return wb

    # Laden außerhalb des Locks: es dauert 0,2–1 s und würde sonst alles blockieren.
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    with _WB_CACHE_LOCK:
        if key in _WB_CACHE:                # ein anderer Thread war schneller
            _WB_CACHE.move_to_end(key)
            return _WB_CACHE[key]
        while len(_WB_CACHE) >= _WB_CACHE_MAX:
            _WB_CACHE.popitem(last=False)   # ältester Zugriff zuerst
        _WB_CACHE[key] = wb
    return wb


def release(data: bytes) -> None:
    """Mappe aus dem Cache werfen — nach dem Einlesen wird sie nicht mehr gebraucht,
    und das Matching danach hat den Speicher lieber selbst."""
    with _WB_CACHE_LOCK:
        _WB_CACHE.pop(_wb_key(data), None)


def _project_name(wb) -> str:
    """Projektname aus den Kopfzeilen des ersten Sheets — längster Text der ersten Zeilen."""
    best = ""
    for ws in wb.worksheets[:3]:
        for r in range(1, 8):
            for c in range(1, 8):
                v = _txt(ws.cell(r, c).value)
                if 8 <= len(v) <= 120 and not _num(v) and len(v) > len(best):
                    best = v
        if best:
            break
    return re.sub(r'\s+', ' ', best)


def iter_scenarios(layout: ExcelLayout) -> list[tuple["SheetLayout", QtyCol, str]]:
    """(Blatt, Mengenspalte, Szenario-Name) für jeden Positionssatz, den parse_excel
    erzeugt. Einzige Quelle dieser Aufzählung.

    Der Export-Dialog hatte vorher seine eigene Regel („Blätter mit >= 2 Mengenspalten")
    und bot damit Szenarien nicht an, die der Parser trotzdem vergibt — die Positionen
    solcher Blätter waren nie auswählbar und bekamen beim Export nie einen Preis.
    """
    active = [sl for sl in layout.sheets if sl.enabled]
    multi  = any(len(sl.roles.active_qty()) > 1 for sl in active)
    out: list[tuple[SheetLayout, QtyCol, str]] = []
    for sl in active:
        for qc in (sl.roles.active_qty() or [QtyCol(col=0)]):
            if not multi:
                out.append((sl, qc, ""))
                continue
            named = (qc.job_name or qc.label or "").strip()
            if named:
                # Ein aus der Datei benannter Wert darf blattübergreifend derselbe
                # sein — dann gehören beide Blätter bewusst zum gleichen Szenario.
                scen = named
            else:
                # Ohne Namen wäre "Menge C" in zwei Blättern derselbe Szenario-Name,
                # und die Auswahl im Export würde beide zusammenwerfen.
                col  = get_column_letter(qc.col) if qc.col else "?"
                scen = f"{sl.name} · Menge {col}" if len(active) > 1 else f"Menge {col}"
            out.append((sl, qc, scen))
    return out


def scenario_names(layout: ExcelLayout) -> list[str]:
    """Auswählbare Szenarien in Dokumentreihenfolge. Leer = nur eines, keine Wahl."""
    names: list[str] = []
    for _, _, scen in iter_scenarios(layout):
        if scen and scen not in names:
            names.append(scen)
    return names


def scenario_name(qc: QtyCol) -> str:
    """Anzeigename eines Szenarios (= Jobname in Easyjob)."""
    return (qc.job_name or qc.label
            or (f"Menge {get_column_letter(qc.col)}" if qc.col else "Menge"))


def layout_fingerprint(data: bytes) -> str:
    """Wiedererkennung derselben Ausschreibungs-Vorlage.

    Bewusst nur aus der Datei abgeleitet, NICHT aus Erkennungsergebnissen: welche
    Rollen-Stichwörter in welcher Spalte im Kopfbereich stehen, plus die Blattnamen.
    Ein früherer Fingerprint über die erkannte Kopfzeile hing an der Erkennung selbst —
    jede Änderung daran machte alle gespeicherten Profile unbrauchbar.

    Ausgefüllte Felder (Bietername, Datum, Preise, Mengen) treffen keines der
    Stichwörter und ändern den Fingerprint deshalb nicht. Dieselbe Vorlage mit anderen
    Zahlen trifft also wieder.
    """
    key = _wb_key(data)
    with _WB_CACHE_LOCK:
        fp = _FP_CACHE.get(key)
        if fp is not None:
            _FP_CACHE.move_to_end(key)
            return fp

    wb = _load(data)
    parts: list[str] = []
    for ws in wb.worksheets:
        found: set[str] = set()
        rows = min(ws.max_row or 1, _MAX_HEADER_SCAN)
        cols = min(ws.max_column or 1, _MAX_COL_SCAN)
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                role = _match_role(ws.cell(r, c).value)
                if role:
                    found.add(f"{role}@{c}")     # Rolle + Spalte = Struktur der Vorlage
        parts.append(f"{ws.title}|{','.join(sorted(found))}")
    fp = hashlib.sha1(chr(10).join(parts).encode("utf-8")).hexdigest()[:16]
    with _WB_CACHE_LOCK:
        while len(_FP_CACHE) >= _FP_CACHE_MAX:
            _FP_CACHE.popitem(last=False)
        _FP_CACHE[key] = fp
    return fp


def _build_probe(ws, layout: SheetLayout, max_col: int,
                 demote_main: bool = False,
                 fml: dict[int, set[int]] | None = None,
                 max_rows: int = _PREVIEW_ROWS) -> SheetProbe:
    """Zeilen klassifizieren und Vorschau für ein Sheet aufbauen.

    fml: Formelzeilen je Spalte (aus formula_rows). Daraus wird gezählt, in wie vielen
    Positionszeilen die Preisspalte eine Formel enthält — der Dialog warnt dann, weil
    dort nicht geschrieben werden darf.
    """
    rows = _classify_sheet(ws, layout, demote_main)
    pos_rows = {r.row for r in rows if r.kind == ROW_POS}
    price_fml = 0
    if fml and layout.roles.price:
        price_fml = len(pos_rows & fml.get(layout.roles.price, set()))
    counts = {
        "pos":  len(pos_rows),
        "job":  sum(1 for r in rows if r.kind == ROW_JOB),
        "main": sum(1 for r in rows if r.kind == ROW_MAIN),
        "note": sum(1 for r in rows if r.kind == ROW_NOTE),
        "price_formulas": price_fml,
    }

    by_row = {r.row: r for r in rows}
    preview: list[PreviewRow] = []
    if layout.header_row:
        preview.append(PreviewRow(
            row=layout.header_row, kind=ROW_HEADER, level=0,
            cells=[_txt(ws.cell(layout.header_row, c).value) for c in range(1, max_col + 1)]))
    # max_rows == 0 heißt: Blatt ist zugeklappt. Gezählt wird trotzdem — sonst wüsste
    # die Leiste nicht, wie viele Zeilen das Blatt hat, und der Umschalter „alle Zeilen"
    # verschwände. Gespart werden nur die Vorschauobjekte (und damit das HTML): bei
    # LOS2 sind das 3484 von 3814 Zellen, die niemand sieht.
    shown, uebrig = 0, 0
    for r in range(layout.first_data_row, (ws.max_row or 0) + 1):
        cells = [_txt(ws.cell(r, c).value) for c in range(1, max_col + 1)]
        if not any(cells):
            continue                      # leere Zeilen zählen gar nicht
        if shown >= max_rows:
            uebrig += 1
            continue
        rr = by_row.get(r)
        preview.append(PreviewRow(row=r, kind=rr.kind if rr else ROW_SKIP,
                                  level=rr.level if rr else 0, cells=cells))
        shown += 1
    counts["shown"]     = shown                  # Datenzeilen in der Vorschau
    counts["truncated"] = uebrig                 # nicht gezeigte Datenzeilen
    counts["rows"]      = shown + uebrig         # Datenzeilen insgesamt
    counts["collapsed"] = 1 if max_rows <= 0 else 0

    ctx, heads = [], []
    if layout.header_row:
        ctx = [[_txt(ws.cell(r, c).value) for c in range(1, max_col + 1)]
               for r in range(max(1, layout.header_row - 4), layout.header_row)]
        heads = [_txt(ws.cell(layout.header_row, c).value) for c in range(1, max_col + 1)]

    return SheetProbe(layout=layout, max_col=max_col, preview=preview, counts=counts,
                      total_rows=ws.max_row or 0, header_context=ctx, header_texts=heads)


def probe_workbook(data: bytes, fmls: dict | None = None,
                   show_all: set[str] | None = None,
                   opened: set[str] | None = None) -> WorkbookProbe:
    """Erkennt Layout + Vorschau für alle Sheets. Rein lesend, keine Seiteneffekte."""
    wb = _load(data)
    probes: list[SheetProbe] = []
    for idx, ws in enumerate(wb.worksheets, start=1):
        layout, max_col = detect_sheet_layout(ws, idx)
        sp = _build_probe(ws, layout, max_col, False, (fmls or {}).get(ws.title),
                          _preview_cap(ws.title, show_all, opened))
        if not sp.counts["pos"]:
            sp.layout.enabled = False       # Deckblatt / Inhalt / Tagessätze
        probes.append(sp)
    return WorkbookProbe(sheets=probes, fingerprint=layout_fingerprint(data),
                         project_name=_project_name(wb))


def preview_workbook(data: bytes, layout: ExcelLayout,
                     fmls: dict | None = None,
                     show_all: set[str] | None = None,
                     opened: set[str] | None = None) -> WorkbookProbe:
    """Vorschau für ein *vorgegebenes* Layout — für den Repreview im Mapping-Dialog
    und für gespeicherte Layout-Profile. Sheets ohne Layout-Eintrag werden erkannt."""
    wb = _load(data)
    n_active = sum(1 for sl in layout.sheets if sl.enabled)
    demote   = n_active > 1 and layout.sheet_mode == SHEET_AS_MAIN
    probes: list[SheetProbe] = []
    for idx, ws in enumerate(wb.worksheets, start=1):
        sl = layout.sheet(ws.title)
        if sl is None:
            sl, max_col = detect_sheet_layout(ws, idx)
            sp = _build_probe(ws, sl, max_col, False, (fmls or {}).get(ws.title),
                              _preview_cap(ws.title, show_all, opened))
            if not sp.counts["pos"]:
                sp.layout.enabled = False
        else:
            sl.sheet_no = idx                       # Reihenfolge kommt aus der Datei
            max_col = _used_max_col(ws, _MAX_HEADER_SCAN + _PREVIEW_ROWS)
            sp_demote = demote and sl.enabled
            # Neu vom Nutzer eingefärbte Mengenspalten haben noch kein Szenario-Label
            if sl.header_row and len(sl.roles.qty) > 1:
                for q in sl.roles.qty:
                    if not q.label:
                        q.label = _qty_label(ws, sl.header_row, q.col)
                    if not q.job_name:
                        q.job_name = q.label
            sp = _build_probe(ws, sl, max_col, sp_demote, (fmls or {}).get(ws.title),
                              _preview_cap(ws.title, show_all, opened))
        probes.append(sp)
    return WorkbookProbe(sheets=probes, fingerprint=layout_fingerprint(data),
                         project_name=_project_name(wb), sheet_mode=layout.sheet_mode)


def merge_profile(probe: WorkbookProbe, saved: dict) -> ExcelLayout:
    """Gespeichertes Profil auf eine frische Probe legen: für jedes Sheet, das im Profil
    vorkommt, gilt die gespeicherte Zuordnung; neue Sheets behalten die Erkennung."""
    stored = layout_from_dict(saved)
    out: list[SheetLayout] = []
    for sp in probe.sheets:
        sl = stored.sheet(sp.layout.name)
        if sl is not None:
            sl.sheet_no = sp.layout.sheet_no
            out.append(sl)
        else:
            out.append(sp.layout)
    return ExcelLayout(sheets=out, fingerprint=probe.fingerprint,
                       label=stored.label, sheet_mode=stored.sheet_mode)


# ─── Parser ─────────────────────────────────────────────────────────────────

class _GroupStack:
    """Gruppenstapel mit zwei Sichten auf dieselbe Hierarchie.

    ``path()``  – nur die äußerste Hauptgruppe. Das ist die Easyjob-Struktur:
                  zwischen Job und Position gibt es genau eine Ebene.
    ``chain()`` – die vollständige Kette der Rohnamen, auch der zu Hinweisen
                  herabgestuften Ebenen. Nur fürs Matching: ``_category_adjustments``
                  fügt den Pfad zu einem String zusammen und sucht Stichwörter darin,
                  je mehr Ebenen also, desto besser („Lichttechnik Rigging Traversen").

    Die Ordnungszahl zählt über alle Ebenen, damit die Reihenfolge stimmt.
    """

    def __init__(self) -> None:
        self.labels:   list[str]  = []   # mit Herkunftskoordinate, für die Anzeige
        self.raw:      list[str]  = []   # ohne Koordinate, fürs Matching
        self.is_hg:    list[bool] = []
        self.counters: list[int]  = []
        self.pos_no:   int        = 0

    def push(self, level: int, label: str, raw: str, is_hg: bool) -> None:
        lvl = max(1, level)
        while len(self.labels) < lvl - 1:          # Ebenenlücke auffüllen
            self.labels.append(""); self.raw.append(""); self.is_hg.append(False)
            self.counters.append(1)
        if len(self.counters) >= lvl:
            self.counters[lvl - 1] += 1
            del self.counters[lvl:]
            del self.labels[lvl - 1:]
            del self.raw[lvl - 1:]
            del self.is_hg[lvl - 1:]
        else:
            self.counters.append(1)
        self.labels.append(label); self.raw.append(raw); self.is_hg.append(is_hg)
        self.pos_no = 0

    def reset(self) -> None:
        self.__init__()

    def path(self) -> list[str]:
        """Die äußerste Hauptgruppe (höchstens eine) — Easyjob hat nur diese Ebene."""
        for lbl, hg in zip(self.labels, self.is_hg):
            if hg and lbl:
                return [lbl]
        return []

    def chain(self) -> list[str]:
        """Alle Ebenen als Rohnamen — Matching-Kontext, nicht Struktur."""
        return [r for r in self.raw if r]

    def next_oz(self) -> str:
        self.pos_no += 1
        return ".".join([f"{n:02d}" for n in self.counters] + [f"{self.pos_no:02d}"])


def _group_label(name: str, sheet_no: int, row: int, qty_suffix: str) -> str:
    """Gruppenname mit Herkunftskoordinate. Zwei Gruppenzeilen können nie dieselbe
    Koordinate haben → Namenskollisionen sind konstruktiv ausgeschlossen (nötig, weil
    ``_import_gaeb_groups`` seine Gruppen allein am Label schlüsselt)."""
    coord = f"{sheet_no}.{row}" + (f".{qty_suffix}" if qty_suffix else "")
    return f"{name} ({coord})"


def parse_excel(data: bytes, layout: ExcelLayout, name: str = "") -> GaebProject:
    """Liest die aktiven Blätter gemäß Layout und baut ein GaebProject.

    Job-Zuordnung, in dieser Reihenfolge zusammengesetzt:
      1. Szenario — jede aktive Mengenspalte ist ein eigener Job
      2. Blatt — nur bei ``sheet_mode == SHEET_AS_JOB``; sonst wird das Blatt eine
         Hauptgruppe und alles landet im selben Job (begrenzt die Job-Anzahl)
      3. Zeilen, die als „Job" gemalt sind, beginnen innerhalb ihres Blatts einen
         weiteren Job

    Das Ergebnis steht in ``job_by_item``; der leere Job-Name = Standard-Job.
    """
    wb = _load(data)
    # Auch Blätter ohne erkannte Kopfzeile werden gelesen, wenn der Nutzer sie
    # angehakt hat — die Daten beginnen dann in Zeile 1. Vorher fielen sie hier
    # lautlos heraus, obwohl die Spalten im Dialog eingefärbt waren.
    active_sheets = [s for s in layout.sheets if s.enabled]
    multi_sheet   = len(active_sheets) > 1
    sheet_as_job  = multi_sheet and layout.sheet_mode == SHEET_AS_JOB
    sheet_as_hg   = multi_sheet and layout.sheet_mode == SHEET_AS_MAIN

    # Mehrere Szenarien = mindestens ein Blatt mit >1 aktiver Mengenspalte. Dann muss
    # die Mengenspalte in Gruppenlabel und item_id, sonst kollidieren die Szenarien.
    # iter_scenarios liefert genau die Kombinationen, die auch der Export-Dialog anbietet.
    scenarios      = iter_scenarios(layout)
    multi_scenario = any(scen for _, _, scen in scenarios)
    by_sheet: dict[str, list[tuple[QtyCol, str]]] = {}
    for sl_, qc_, scen_ in scenarios:
        by_sheet.setdefault(sl_.name, []).append((qc_, scen_))

    items:    list[GaebItem]   = []
    remarks:  list[GaebRemark] = []
    job_by:   dict[str, str]   = {}
    scen_by:  dict[str, str]   = {}

    for sl in active_sheets:
        ws = wb[sl.name]
        # Blatt = Hauptgruppe → Gruppen darin werden Hinweise (wie in der Vorschau)
        rows = _classify_sheet(ws, sl, demote_main=sheet_as_hg)
        for qc, scen in by_sheet.get(sl.name, []):
            suffix = get_column_letter(qc.col) if (multi_scenario and qc.col) else ""

            # Job-Präfix aus Szenario und (optional) Blatt
            prefix_parts = [p for p in (scen, sl.name if sheet_as_job else "") if p]
            job_prefix   = " · ".join(prefix_parts)
            cur_job      = job_prefix

            # Blatt als Hauptgruppe (Modus main) bzw. als Kontextebene darüber
            # (Modus keep). Bei mehreren Szenarien braucht auch dieses Label die
            # Mengenspalte, sonst verschmelzen die Szenarien.
            sheet_label = ""
            if multi_sheet and not sheet_as_job:
                sheet_label = f"{sl.name} ({suffix})" if suffix else sl.name

            def _root(job: str) -> list[str]:
                """Wurzel des category_path: Job (falls vorhanden) und/oder Blatt.
                Das Blatt steht hier und nicht im Gruppenstapel — sonst würde es von
                der Hauptgruppe verdrängt, die den Stapel auf eine Ebene reduziert."""
                return ([job] if job else []) + ([sheet_label] if sheet_label else [])

            # Matching-Wurzel ohne Koordinaten: Szenario/Blattname sind echte Stichwörter
            root_raw = [t for t in (scen, sl.name) if t]

            stack   = _GroupStack()
            pending: list[GaebRemark] = []

            for r in rows:
                if r.kind == ROW_JOB:
                    # Neuer Job innerhalb des Blatts — Gruppenstapel beginnt von vorn
                    label   = _group_label(r.label, sl.sheet_no, r.row, suffix)
                    cur_job = f"{job_prefix} · {label}" if job_prefix else label
                    stack.reset()
                    for rem in pending:
                        rem.category_path = _root(cur_job) + stack.path()
                        remarks.append(rem)
                    pending = []
                    continue

                # Hauptgruppen UND herabgestufte Gruppen gehören in den Stapel: die
                # einen als Struktur, die anderen als Matching-Kontext.
                if r.kind in (ROW_MAIN, ROW_GROUP) or r.was_group:
                    stack.push(r.level,
                               _group_label(r.label, sl.sheet_no, r.row, suffix),
                               r.label, r.kind in (ROW_MAIN, ROW_GROUP))
                    if r.was_group:
                        # Zwischenüberschrift → Hinweis vor der nächsten Position
                        pending.append(GaebRemark(title=r.label, long_text=r.long,
                                                  category_path=[], images=[], next_item_id=""))
                    for rem in pending:
                        if not rem.category_path:
                            rem.category_path = _root(cur_job) + stack.path()
                    continue

                path = _root(cur_job) + stack.path()

                if r.kind == ROW_NOTE:
                    pending.append(GaebRemark(title=r.label, long_text=r.long,
                                              category_path=path, images=[], next_item_id=""))
                    continue

                # ── Position ──
                item_id = f"{sl.sheet_no}-{r.row}" + (f"-{qc.col}" if qc.col else "")
                qty = r.qty.get(qc.col) if qc.col else None
                oz  = r.oz or stack.next_oz()

                long_parts = [t for t in (r.long, r.ref) if t]
                items.append(GaebItem(
                    item_id=item_id,
                    rno_part=r.row,
                    oz=oz,
                    description=r.label,
                    long_text=_NL.join(long_parts),
                    qty=float(qty) if qty is not None else 0.0,
                    unit=r.unit,
                    category_path=path,
                    is_alt=bool(_ALT_RE.match(r.flag.strip())) or
                           bool(re.match(r'^al?ternativ', r.label, re.IGNORECASE)),
                    is_eventual=bool(_EVENT_RE.match(r.flag.strip())) or
                                bool(re.match(r'^(bedarfs|eventual)', r.label, re.IGNORECASE)),
                    src_ref=f"{sl.name}!{r.row}",
                    ref_text=r.ref,
                    match_path=root_raw + stack.chain(),
                ))
                job_by[item_id]  = cur_job
                scen_by[item_id] = scen
                for rem in pending:
                    rem.category_path = path
                    rem.next_item_id  = item_id
                    remarks.append(rem)
                pending = []

            for rem in pending:                    # Hinweise am Blatt-Ende
                rem.category_path = _root(cur_job) + stack.path()
                remarks.append(rem)

    return GaebProject(name=name or layout.label or "Excel-LV", label="",
                       phase="XLSX", date="", currency="EUR",
                       items=items, remarks=remarks, preliminaries=[],
                       job_by_item=job_by, scenario_by_item=scen_by)
