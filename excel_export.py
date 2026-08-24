"""Preis-Rückschreibung in die Original-Excel — das Gegenstück zum GAEB-X84-Export.

Die hochgeladene Datei wird unverändert weiterverwendet und nur in den Preisspalten
befüllt: Formatierung, Formeln und alle Blätter bleiben, damit der Auftraggeber sein
eigenes Preisblatt zurückbekommt und nicht eine neu erzeugte Datei.
"""
import io
import logging
import zipfile
from dataclasses import dataclass

import openpyxl
from openpyxl.utils import get_column_letter

import excel_parser as _xl
from gaeb_parser import GaebItem, GaebProject


@dataclass
class WriteResult:
    data:    bytes
    written: int          # Positionen mit Preis
    skipped: int          # Positionen ohne Preisspalte oder ohne Zeilenbezug
    formulas: int         # Gesamtpreis-Zellen, die als Formel stehen blieben
    notes:   list[str]
    suffix:  str = ".xlsx"   # Endung der Ausgabe — .xlsm behält seine Makros


def _parse_src_ref(ref: str) -> tuple[str, int]:
    """"Los 1 Technik!18" → ("Los 1 Technik", 18). Sheet-Namen dürfen "!" enthalten,
    deshalb wird am *letzten* Trenner geteilt."""
    if "!" not in ref:
        return "", 0
    sheet, _, row = ref.rpartition("!")
    return sheet, int(row) if row.isdigit() else 0


def _total_col(roles: _xl.ColumnRoles) -> int:
    """Die zugeordnete Gesamtpreis-Spalte, sonst 0.

    Vorher wurde sie geraten („direkt rechts vom Einzelpreis, wenn dort keine andere
    Rolle sitzt"). ``role_of`` gibt aber für JEDE nicht zugeordnete Spalte
    ROLE_IGNORE zurück — auch für eine Bemerkungsspalte des Kunden oder eine, die es
    gar nicht gibt. Der Export hat dort Zahlen hineingeschrieben. Jetzt wird nur
    befüllt, was im Mapping-Dialog ausdrücklich als Gesamtpreis markiert ist.
    """
    return roles.total


def _is_macro_workbook(source: bytes) -> bool:
    """Enthält die Mappe ein VBA-Projekt? Dann muss sie als .xlsm zurückkommen."""
    try:
        with zipfile.ZipFile(io.BytesIO(source)) as z:
            return any(n.endswith("vbaProject.bin") for n in z.namelist())
    except (zipfile.BadZipFile, OSError):
        return False


def write_prices(source: bytes, layout: _xl.ExcelLayout, project: GaebProject,
                 prices: dict[str, float]) -> WriteResult:
    """Schreibt Einheitspreise in die Original-Excel.

    ``prices`` bildet item_id → Einheitspreis. Die Zielzelle ergibt sich aus
    ``GaebItem.src_ref`` (Sheet + Zeile) und der Preisspalte des Sheet-Layouts.
    Der Gesamtpreis wird nur gesetzt, wenn dort keine Formel steht — sonst rechnet
    Excel selbst weiter und die Datei bleibt für den Auftraggeber prüfbar.
    """
    # data_only=False: Formeln und Formatierung der übrigen Zellen bleiben erhalten.
    # keep_vba: bei einer Makro-Mappe (.xlsm) würde openpyxl das VBA-Projekt sonst
    # beim Speichern wegwerfen — Ausschreibungs-Preisblätter tragen oft Makros für
    # Prüfungen und Berechnungen, die der Auftraggeber zurückerwartet.
    makros = _is_macro_workbook(source)
    wb = openpyxl.load_workbook(io.BytesIO(source), data_only=False, keep_vba=makros)

    by_sheet: dict[str, _xl.SheetLayout] = {s.name: s for s in layout.sheets}
    items_by_id: dict[str, GaebItem] = {i.item_id: i for i in project.items}

    written = skipped = formulas = 0
    notes: list[str] = []
    missing_price_col: set[str] = set()
    formula_cols: dict[tuple[str, int], int] = {}   # (Blatt, Spalte) -> Anzahl Formeln

    # Mehrere Szenarien schreiben in dieselbe Preisspalte — der zuletzt gewonnene Preis
    # würde den ersten überschreiben. Je Zelle den höchsten Preis behalten und melden.
    seen: dict[tuple[str, int], float] = {}

    for item_id, ep in prices.items():
        item = items_by_id.get(item_id)
        if item is None or not item.src_ref:
            skipped += 1
            continue
        sheet_name, row = _parse_src_ref(item.src_ref)
        sl = by_sheet.get(sheet_name)
        if sl is None or not row:
            skipped += 1
            continue
        if not sl.roles.price:
            missing_price_col.add(sheet_name)
            skipped += 1
            continue
        if sheet_name not in wb.sheetnames:
            skipped += 1
            continue

        key_row = (sheet_name, row)
        if key_row in seen:
            if ep <= seen[key_row]:
                continue                     # niedrigeren Preis nicht überschreiben
            notes.append(f"{sheet_name}!{row}: mehrere Szenarien — höchster EP übernommen")

        ws   = wb[sheet_name]
        cell = ws.cell(row, sl.roles.price)

        # Eine Formel NIE überschreiben. In manchen Preisblättern ist ausgerechnet die
        # Spalte, die "Einzelpreis" heißt, selbst berechnet (LOS1 GOE: E = C + D, also
        # Material + Personal). Sie zu überschreiben würde die Rechenlogik der Datei
        # zerstören — schlimmer als ein fehlender Preis. Also melden, damit im Mapping
        # die Eingabespalte als Einzelpreis markiert wird.
        if isinstance(cell.value, str) and cell.value.startswith("="):
            k = (sheet_name, sl.roles.price)
            formula_cols[k] = formula_cols.get(k, 0) + 1
            skipped += 1
            continue

        seen[key_row] = ep
        cell.value = round(float(ep), 2)
        written += 1

        tcol = _total_col(sl.roles)
        if tcol:
            tcell = ws.cell(row, tcol)
            if isinstance(tcell.value, str) and tcell.value.startswith("="):
                formulas += 1                # Excel rechnet selbst — nicht anfassen
            else:
                tcell.value = round(float(ep) * float(item.qty or 0), 2)

    for name in sorted(missing_price_col):
        notes.append(f"Blatt „{name}“: keine Einzelpreis-Spalte zugeordnet — "
                     f"Preise nicht geschrieben")
    for (name, col), n in sorted(formula_cols.items()):
        notes.append(
            f"Blatt „{name}“, Spalte {get_column_letter(col)}: {n} Zellen enthalten "
            f"Formeln und wurden NICHT überschrieben — diese Spalte berechnet die Datei "
            f"selbst. Im Mapping bitte die Eingabespalte als Einzelpreis markieren "
            f"(z.B. den Material- bzw. Personalanteil).")

    out = io.BytesIO()
    wb.save(out)
    logging.info("excel_export: %d Preise geschrieben, %d übersprungen, "
                 "%d Formeln belassen", written, skipped, formulas)
    return WriteResult(data=out.getvalue(), suffix=".xlsm" if makros else ".xlsx",
                       written=written, skipped=skipped, formulas=formulas, notes=notes)
