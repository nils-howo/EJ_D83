"""Durchläuft die Excel-Export-Route komplett — ohne Easyjob-Datenbank.

Diese Route braucht im Betrieb eine EJ-DB-Verbindung und wurde deshalb nie ausgeführt;
ein NameError in ihrem Rumpf fiel erst beim Anwender auf. Hier wird nur die
Kostenabfrage ersetzt, alles andere läuft echt: Layout laden, Positionen bauen, Preise
schreiben, Dateiname, Hinweise, Download.

Zusätzlich ein statischer Durchlauf über alle Module, der Namen findet, die gelesen
aber nirgends gebunden werden — genau die Fehlerklasse von oben.

    .venv/Scripts/python.exe tests/test_export_route.py
"""
import sys

# Konsole auf UTF-8: sonst stirbt schon ein "→" im print an cp1252 und der Test
# bricht mitten drin ab, ohne dass eine Prüfung fehlgeschlagen wäre.
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import ast
import builtins
import io
import json
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
os.chdir(os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from fastapi import FastAPI
from starlette.middleware.sessions import SessionMiddleware
from starlette.testclient import TestClient

import db as _db
import excel_parser as _xl
import routes.projects as pj

_fails: list[str] = []
_oks = 0


def check(cond, msg: str) -> None:
    global _oks
    if cond:
        _oks += 1
    else:
        _fails.append(msg)
        print(f"  FAIL  {msg}")


# ── 1. Statisch: gelesene, aber nirgends gebundene Namen ────────────────────

def _bound_in(fn) -> set[str]:
    """Namen, die diese Funktion selbst bindet (Parameter, Zuweisungen, Imports …).
    Nur die eigene Ebene — verschachtelte Funktionen bringen ihre eigenen mit."""
    out: set[str] = set()
    a = fn.args
    out.update(x.arg for x in a.posonlyargs + a.args + a.kwonlyargs)
    for extra in (a.vararg, a.kwarg):
        if extra:
            out.add(extra.arg)

    def walk(node):
        for child in ast.iter_child_nodes(node):
            if isinstance(child, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
                out.add(child.name)      # der Name ja, der Rumpf ist eine eigene Ebene
                continue
            if isinstance(child, ast.Name) and isinstance(child.ctx, (ast.Store, ast.Del)):
                out.add(child.id)
            elif isinstance(child, (ast.Import, ast.ImportFrom)):
                out.update((n.asname or n.name).split(".")[0] for n in child.names)
            elif isinstance(child, ast.ExceptHandler) and child.name:
                out.add(child.name)
            elif isinstance(child, (ast.Global, ast.Nonlocal)):
                out.update(child.names)
            elif isinstance(child, ast.Lambda):
                # Lambda-Parameter gelten in ihrem Rumpf; grob dem Umfeld zurechnen,
                # sonst wäre jeder Lambda-Parameter ein Falschtreffer.
                la = child.args
                out.update(x.arg for x in la.posonlyargs + la.args + la.kwonlyargs)
                for extra2 in (la.vararg, la.kwarg):
                    if extra2:
                        out.add(extra2.arg)
            walk(child)

    walk(fn)
    return out


def _loads_in(fn) -> list[tuple[str, int]]:
    """Gelesene Namen auf dieser Ebene — ohne die Rümpfe verschachtelter Funktionen."""
    out: list[tuple[str, int]] = []

    def walk(node):
        for child in ast.iter_child_nodes(node):
            if isinstance(child, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
                continue
            if isinstance(child, ast.Name) and isinstance(child.ctx, ast.Load):
                out.append((child.id, child.lineno))
            walk(child)

    walk(fn)
    return out


def undefined_names(path: str) -> list[str]:
    """Namen, die gelesen werden, obwohl sie nirgends im Sichtbarkeitsbereich gebunden
    sind — weder lokal, noch in einer umgebenden Funktion, noch als Modul-Global oder
    Builtin. Findet den Fall "Zeile beim Refactoring gelöscht, Verwendung stehen
    geblieben". Findet NICHT: UnboundLocalError durch `x |= ...` in einer Closure."""
    tree = ast.parse(open(path, encoding="utf-8").read())

    module: set[str] = set(dir(builtins)) | {"__file__", "__name__", "__doc__"}
    for node in ast.walk(tree):
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
            module.add(node.name)
        elif isinstance(node, ast.Name) and isinstance(node.ctx, ast.Store):
            module.add(node.id)
        elif isinstance(node, (ast.Import, ast.ImportFrom)):
            module.update((a.asname or a.name).split(".")[0] for a in node.names)

    bad: list[str] = []

    def visit(node, outer: set[str]) -> None:
        for child in ast.iter_child_nodes(node):
            if isinstance(child, (ast.FunctionDef, ast.AsyncFunctionDef)):
                scope = outer | _bound_in(child)
                for name, line in _loads_in(child):
                    if name not in scope:
                        bad.append(
                            f"{path}:{line} {child.name}() liest undefiniertes {name!r}")
                visit(child, scope)      # verschachtelte Funktionen erben den Bereich
            else:
                visit(child, outer)

    visit(tree, module)
    return sorted(set(bad))


MODULES = ("routes/projects.py", "routes/import_.py", "routes/auth.py",
           "routes/admin.py", "routes/matching.py", "routes/d83.py",
           "excel_parser.py", "excel_export.py", "gaeb_parser.py",
           "server.py", "db.py", "state.py", "matcher.py")


def test_static():
    print("\n=== Statischer Check auf undefinierte Namen")
    total = 0
    for mod in MODULES:
        bad = undefined_names(mod)
        total += len(bad)
        for b in bad:
            print(f"  {b}")
        check(not bad, f"{mod}: {len(bad)} undefinierte Namen")
    print(f"  {len(MODULES)} Module, {total} Funde")


# ── 2. Die Route selbst, mit gefälschter Kostenbasis ────────────────────────

# (Datei, Blatt, zwei Mengenspalten, landen Preise wirklich in Zellen?)
#   PAG: Spalte F ist eine Eingabezelle          -> Preise werden geschrieben
#   GOE: die als "Einzelpreis" erkannte Spalte E ist eine Formel (=C+D)
#        -> darf NICHT überschrieben werden, der Export muss es melden
CASES = [
    ("PAG_NOC2026_LV Veranstaltungstechnik_V1.xlsx", "01_Material", [5, 8], True),
    ("LOS1_GOE Ausschreibung_Technik 2026_Preisblatt_V2.1.xlsx",
     "Preisblatt Technik", [8, 12], False),
]


def _fake_costs(gid_by_item, cost_per_group):
    """Ersetzt _export_costs — sonst bräuchte der Test eine Easyjob-Datenbank."""
    def _inner(ss, project_id, proj):
        return {
            "job_ids": [900, 901], "bookings": [],
            "group_by_item":   gid_by_item,
            "qty_by_item":     {},
            "group_art_cost":  cost_per_group,
            "group_pers_cost": {},
            "oz_to_group":     {},
            "learned":         [],
            "items_per_group": {g: [i] for i, g in gid_by_item.items()},
            "item_cost":       {},
            "shared_groups":   0,
        }
    return _inner


class _StubSession:
    """export_dialog legt keine Session an, die echte bräuchte einen Easyjob-Login.
    Die Route prüft nur, ob ej_db_conn gesetzt ist."""

    def __init__(self):
        self.ej_db_conn = "stub"
        self.pending_export = None


def _app():
    app = FastAPI()
    app.include_router(pj.router)
    app.add_middleware(SessionMiddleware, secret_key="t" * 40, session_cookie="t")
    return app


def _run_case(fname, sheet_name, qty_cols, expect_written):
    print(f"\n=== {fname[:46]}")
    src = open(os.path.join("infos", fname), "rb").read()

    layout = _xl.probe_workbook(src).layout
    sl = layout.sheet(sheet_name)
    # Zwei Szenarien erzwingen, damit der Dialog die Auswahl anbietet
    sl.roles.qty = [_xl.QtyCol(col=c, label=f"Szenario {c}", job_name=f"Szenario {c}")
                    for c in qty_cols]
    project = _xl.parse_excel(src, layout)
    gid_by_item = {it.item_id: 1000 + n for n, it in enumerate(project.items)}
    cost_per_group = {g: 200.0 for g in gid_by_item.values()}

    # Die Zuordnung hängt am Projekt — genau wie der echte Import sie schreibt.
    pid = _db.save_project(name="Export-Test", ej_project_id=4711, gaeb_name=fname,
                           item_count=0, booking_count=0, gaeb_bytes=src,
                           ej_job_ids="900,901", ej_project_number="26-TEST",
                           source_kind="excel",
                           source_layout_json=json.dumps(_xl.layout_to_dict(layout)))

    # Sabotiertes globales Profil unter demselben Fingerprint: der Export DARF es
    # nicht anfassen, sonst schreibt ein späterer Import fremde Spalten in dieses
    # Angebot. Preisspalte hier absichtlich verschoben.
    fremd = _xl.layout_to_dict(layout)
    for sh in fremd["sheets"]:
        if sh["roles"]["price"]:
            sh["roles"]["price"] = 1          # Spalte A — würde die Pos-Nr überschreiben
    _db.save_excel_layout(layout.fingerprint, "FREMD", fremd)

    ss = _StubSession()
    orig_costs, orig_sess = pj._export_costs, pj.get_session
    pj._export_costs = _fake_costs(gid_by_item, cost_per_group)
    pj.get_session = lambda _s: ss
    try:
        with TestClient(_app()) as c:
            r = c.get(f"/api/projects/{pid}/export-dialog")
            check(r.status_code == 200, f"Dialog HTTP {r.status_code}")
            # Erwartung aus derselben Quelle, die auch der Parser benutzt — eine
            # feste Zahl hier hatte genau den Fehler zementiert, den Fix 4 behebt
            # (Blätter mit einer Mengenspalte waren nie auswählbar).
            erwartet = _xl.scenario_names(layout)
            n_radios = r.text.count('name="scenario"')
            check(n_radios == len(erwartet),
                  f"{len(erwartet)} Szenario-Radios erwartet ({erwartet}), sind {n_radios}")
            check(len(erwartet) == len(set(erwartet)),
                  f"Szenario-Namen müssen eindeutig sein: {erwartet}")
            check("export-excel" in r.text, "Dialog muss auf export-excel zeigen")

            scen = f"Szenario {qty_cols[1]}"
            check(scen in erwartet, f"{scen!r} muss anwählbar sein, sind {erwartet}")
            r = c.post(f"/api/projects/{pid}/export-excel", data={"scenario": scen})
            check(r.status_code == 200, f"Export HTTP {r.status_code}: {r.text[:200]}")
            if r.status_code != 200:
                return

            pe = ss.pending_export
            print(f"  Dateiname: …{pe['name'][-42:]}")
            check(pe["name"].endswith(".xlsx"), f"Endung: {pe['name']}")
            suffix = pe["name"].split("_Angebot", 1)[1]
            check(" " not in suffix, f"Szenario-Teil bereinigt: {suffix!r}")
            check(pe["project_id"] == pid, "pending_export gehört zum Projekt")
            check("Positionen mit Preis" in r.text, "Anzahl-Hinweis fehlt")
            check("Szenario" in r.text, "Szenario-Hinweis fehlt")

            d = c.get(f"/api/projects/{pid}/export-d84/download")
            check(d.status_code == 200, f"Download HTTP {d.status_code}")
            neu = openpyxl.load_workbook(io.BytesIO(d.content), data_only=False)
            alt = openpyxl.load_workbook(io.BytesIO(src), data_only=False)
            check(neu.sheetnames == alt.sheetnames, "Blätter unverändert")

            ws_neu, ws_alt = neu[sheet_name], alt[sheet_name]
            pcol = sl.roles.price
            row = min(int(i.src_ref.rsplit("!", 1)[1]) for i in project.items
                      if i.src_ref.startswith(sheet_name + "!"))
            vor, nach = ws_alt.cell(row, pcol).value, ws_neu.cell(row, pcol).value
            print(f"  Spalte {pcol}, Zeile {row}: {vor!r} -> {nach!r}")
            if expect_written:
                check(isinstance(nach, (int, float)) and nach > 0,
                      f"Preis erwartet, ist {nach!r}")
                # "Formeln" allein reicht nicht: der harmlose Hinweis über belassene
                # Gesamtpreis-Formeln enthält das Wort auch.
                check("NICHT" not in r.text,
                      "keine Meldung über eine berechnete Preisspalte erwartet")
            else:
                check(nach == vor, f"Formel wurde überschrieben: {nach!r}")
                check("Formeln" in r.text and "NICHT" in r.text,
                      "Export muss die berechnete Spalte melden")
                print("  -> berechnete Spalte gemeldet statt überschrieben")

            # Das fremde Profil darf nicht durchgeschlagen haben: Spalte A (Pos-Nr)
            # muss unverändert sein.
            check(ws_neu.cell(row, 1).value == ws_alt.cell(row, 1).value,
                  "fremdes Layout-Profil hat die Pos-Nr-Spalte überschrieben")

            # Mehrere Szenarien -> ohne Auswahl muss der Export ablehnen, statt sie
            # gegenseitig in dieselben Zellen schreiben zu lassen.
            r2 = c.post(f"/api/projects/{pid}/export-excel", data={})
            print(f"  ohne Szenario-Auswahl -> HTTP {r2.status_code}")
            check(r2.status_code == 400,
                  f"400 erwartet (Szenario-Pflicht), ist {r2.status_code}")
    finally:
        pj._export_costs, pj.get_session = orig_costs, orig_sess
        with _db.get_conn() as cn:
            cn.execute("DELETE FROM projects WHERE id=?", (pid,))
            cn.execute("DELETE FROM excel_layouts WHERE fingerprint=?",
                       (layout.fingerprint,))


def test_route():
    for case in CASES:
        _run_case(*case)


def test_layout_fehlt():
    print(f"\n=== Excel-Projekt ohne gespeicherte Spalten-Zuordnung")
    src = open(os.path.join("infos", CASES[0][0]), "rb").read()
    pid = _db.save_project(name="Export-Test", ej_project_id=1, gaeb_name="x.xlsx",
                           item_count=0, booking_count=0, gaeb_bytes=src,
                           ej_job_ids="1", source_kind="excel")   # kein Layout!
    orig = pj.get_session
    pj.get_session = lambda _s: _StubSession()
    try:
        with TestClient(_app()) as c:
            r = c.post(f"/api/projects/{pid}/export-excel", data={})
            print(f"  HTTP {r.status_code}")
            check(r.status_code == 400, f"400 erwartet, ist {r.status_code}")
            check("Zuordnung" in r.text, f"Meldung soll die Ursache nennen: {r.text[:120]}")
    finally:
        pj.get_session = orig
        with _db.get_conn() as cn:
            cn.execute("DELETE FROM projects WHERE id=?", (pid,))


def test_wrong_source():
    print("\n=== GAEB-Projekt darf die Excel-Route nicht benutzen")
    pid = _db.save_project(name="GAEB-Test", ej_project_id=1, gaeb_name="x.x83",
                           item_count=0, booking_count=0, gaeb_bytes=b"<x/>",
                           ej_job_ids="1", source_kind="gaeb")
    orig = pj.get_session
    pj.get_session = lambda _s: _StubSession()
    try:
        with TestClient(_app()) as c:
            r = c.post(f"/api/projects/{pid}/export-excel", data={})
            print(f"  HTTP {r.status_code}")
            check(r.status_code == 400, f"400 erwartet, ist {r.status_code}")
    finally:
        pj.get_session = orig
        with _db.get_conn() as cn:
            cn.execute("DELETE FROM projects WHERE id=?", (pid,))


# Von pytest gesehen: check() sammelt nur, damit ein Lauf ALLE Fehlschläge zeigt.
# Ohne diesen Abschluss meldet pytest jede Testfunktion als PASSED, auch wenn jede
# einzelne Prüfung fehlgeschlagen ist.
def test_zz_alle_pruefungen_ok():
    assert not _fails, f"{len(_fails)} Prüfung(en) fehlgeschlagen: " + "; ".join(_fails)


if __name__ == "__main__":
    test_static()
    test_route()
    test_layout_fehlt()
    test_wrong_source()
    print("\n" + "=" * 60)
    print(f"{_oks} Prüfungen ok, {len(_fails)} fehlgeschlagen")
    if _fails:
        for f in _fails:
            print(f"  - {f}")
        sys.exit(1)
