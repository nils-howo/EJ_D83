"""Regressionstest für den Export der Personalplanung (PDF und Excel).

Die Beilage geht zum Kunden — deshalb prüft dieser Test vor allem die eine Sache, die
teuer ist, wenn sie schiefgeht: **in der Kundenversion steht kein Preis.** Der Rest
sind Rahmenbedingungen (eine Seite je Monat, Formeln statt Zahlen in der Excel,
sprechende Dateinamen).

    .venv/Scripts/python.exe tests/test_crew_export.py
"""
import sys

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import asyncio
import base64
import io as _io
import os
import re
import tempfile
import zlib

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

_TMP_DB = os.path.join(tempfile.mkdtemp(prefix="crewexp_"), "test.db")
os.environ["DB_PATH"] = _TMP_DB

from crew_plan import CrewPlan, Phase                    # noqa: E402
from crew_pdf import build_pdf                           # noqa: E402
from crew_xlsx import build_xlsx                         # noqa: E402

_fails: list[str] = []


def check(name, ist, soll):
    ok = ist == soll
    print(("  ✓ " if ok else "  ✗ ") + name + ": " + repr(ist)
          + ("" if ok else "  ERWARTET " + repr(soll)))
    if not ok:
        _fails.append(name)


_LOOP = asyncio.new_event_loop()


def run(coro):
    return _LOOP.run_until_complete(coro)


def _plan(tage=31):
    p = CrewPlan(date_from="2026-03-09", date_to="2026-04-08")
    p.positions = ["p1"]
    p.phases = [Phase(name="Aufbau", day_from="2026-03-09", day_to="2026-03-14"),
                Phase(name="Veranstaltung", day_from="2026-03-15",
                      day_to="2026-04-05"),
                Phase(name="Abbau", day_from="2026-04-06", day_to="2026-04-08")]
    mk = p.add_custom_menu("Projektleitung")
    for name, ts, n, grp in [("Projektleiter", 850.0, 20, mk),
                             ("Ton-Operator", 600.0, 12, "")]:
        r = p.add_row(name, 500 + len(p.rows), group_key=grp)
        r.tagessatz, r.hotel_naechte, r.rk_anzahl = ts, n - 2, 2
        for k in p.day_keys()[:n]:
            p.set_cell(r.id, k, 2 if "Ton" in name else 1)
        p.assign_days(r.id, p.day_keys()[0], p.day_keys()[n - 1], "p1")
    return p


# ─── 1. PDF ──────────────────────────────────────────────────────────────────
print("── 1. PDF ──")
plan = _plan()

pdf_kalk = build_pdf(plan, "Schneider Electric @ HMI 2026", kalkulation=True)
pdf_kunde = build_pdf(plan, "Schneider Electric @ HMI 2026", kalkulation=False)
check("PDF entsteht", pdf_kalk[:5], b"%PDF-")
check("beide Varianten", (len(pdf_kalk) > 1000, len(pdf_kunde) > 1000), (True, True))


_STREAM = re.compile(rb"stream\r?\n(.*?)endstream", re.S)
_LITERAL = re.compile(r"\(((?:[^()\\]|\\.)*)\)")


def _text(daten: bytes) -> str:
    """Den sichtbaren Text aus dem PDF ziehen — nur mit der Standardbibliothek.

    reportlab legt den Seiteninhalt als Flate-komprimierten Strom ab; darin stehen die
    Zeichenketten als PDF-Literale in Klammern. Das reicht, um zu prüfen, was auf dem
    Blatt landet, und macht die Prüfung von keinem zusätzlichen Paket abhängig.
    """
    stuecke = []
    for m in _STREAM.finditer(daten):
        roh = m.group(1).strip()
        if roh.endswith(b"~>"):             # ASCII85 vor Flate — reportlabs Vorgabe
            roh = roh[:-2]
        try:
            entpackt = zlib.decompress(base64.a85decode(roh))
        except Exception:                   # noqa: BLE001 — kein Flate-Strom
            continue
        # Nur Seiteninhalt, keine Bilder: die Pixel des Briefbogens lassen sich
        # genauso entpacken, und in ihnen sieht zufällig alles wie Text aus — auch
        # ein Euro-Zeichen. „BT" beginnt einen Textblock und kommt nur im Inhalt vor.
        if b"BT" in entpackt:
            stuecke.append(entpackt)
    inhalt = b"\n".join(stuecke).decode("latin-1")
    # Oktal-Escapes auflösen und aus WinAnsi zurückübersetzen: das Euro-Zeichen steht
    # dort auf 0x80 und käme sonst als Steuerzeichen an statt als €.
    worte = [re.sub(r"\\([0-7]{3})", lambda m2: chr(int(m2.group(1), 8)), w)
             for w in _LITERAL.findall(inhalt)]
    text = "\n".join(worte)
    return text.encode("latin-1", "replace").decode("cp1252", "replace")


def _seiten(daten: bytes) -> int:
    """Seitenzahl aus den Objekttypen — /Pages ist der Baumknoten, nicht die Seite."""
    return daten.count(b"/Type /Page") - daten.count(b"/Type /Pages")


t_kalk, t_kunde = _text(pdf_kalk), _text(pdf_kunde)
check("Text lesbar", len(t_kalk) > 100, True)

# Das Wesentliche: die Kundenversion trägt keine Beträge. Eine Datei mit dem falschen
# Inhalt fällt beim Verschicken niemandem auf.
check("Kalkulation zeigt Beträge", "€" in t_kalk, True)
check("Kundenversion ohne Beträge", "€" in t_kunde, False)
check("Kundenversion ohne Tagessatz", "850" in t_kunde, False)
check("Kalkulation mit Tagessatz", "850" in t_kalk, True)
# Beide zeigen dieselbe Besetzung — es ist eine Liste, nicht zwei.
for _name in ("Projektleiter", "Ton-Operator", "PROJEKTLEITUNG"):
    check("„" + _name + "“ in beiden", (_name in t_kalk, _name in t_kunde),
          (True, True))
check("Phasen im Kopf", "VERANSTALTUNG" in t_kunde, True)
# Die Kundenfassung trägt keinen Hinweis darauf, dass ihr etwas fehlt — sie ist die
# Besetzungsliste, nicht die gekürzte Fassung von etwas anderem.
check("kein Hinweis auf fehlende Preise",
      ("bersicht" in t_kunde, "ohne Preise" in t_kunde), (False, False))
# Der Monatsname wird auf die Breite seines Blocks gekürzt. Fängt der Zeitraum am
# 31. an, ist der erste Monat einen Tag breit — der volle Name lief bisher quer
# über den nächsten, und beide standen übereinander.
_31 = CrewPlan(date_from="2026-03-31", date_to="2026-04-25")
_31.positions = ["a"]
_r31 = _31.add_row("Techniker", 501)
for _k in _31.day_keys()[:12]:
    _31.set_cell(_r31.id, _k, 1)
_31.assign_days(_r31.id, _31.day_keys()[0], _31.day_keys()[11], "a")
_t31 = _text(build_pdf(_31, "Start am 31."))
check("kurzer Monat gekürzt", ("März 2026" in _t31, "März" in _t31), (False, True))
check("der nächste voll", "April 2026" in _t31, True)
# Der Briefbogen kommt als Bild aufs Blatt, nicht als Text — sonst müsste die Adresse
# hier gepflegt werden und liefe der echten irgendwann hinterher.
check("Briefbogen ist ein Bild", "Leonberg" in t_kunde, False)

# Ein Monat muss auf eine Seite passen: ein Umbruch mitten in der Veranstaltung macht
# die Beilage unlesbar.
check("31 Tage auf einer Seite", _seiten(pdf_kalk), 1)
lang = CrewPlan(date_from="2026-03-01", date_to="2026-05-31")
lang.positions = ["p1"]
_r = lang.add_row("Techniker", 501)
lang.set_cell(_r.id, "2026-03-01", 1)
check("92 Tage brauchen mehr", _seiten(build_pdf(lang, "Lang")) > 1, True)

# Zeilen ohne einen einzigen Manntag gehören nicht auf das Blatt: in der Matrix
# stehen sie als Vorschlag aus dem Matching, ausgedruckt sind sie eine Liste von
# Leuten, die nicht kommen.
_ohne = _plan()
_leerzeile = _ohne.add_row("Niemals dabei", 999)
_ohne.rows[-1].tagessatz = 700.0
_t_ohne = _text(build_pdf(_ohne, "Mit Leerzeile"))
check("unbesetzte Zeile fehlt im PDF", "Niemals dabei" in _t_ohne, False)
check("besetzte Zeilen bleiben", "Ton-Operator" in _t_ohne, True)
# Ein Abschnitt, in dem danach nichts mehr steht, fällt gleich mit weg.
_a = CrewPlan(date_from="2026-03-09", date_to="2026-03-12")
_a.positions = ["p1"]
_k = _a.add_custom_menu("Leerer Abschnitt")
_a.add_row("Unbesetzt", 998, group_key=_k)
check("leerer Abschnitt fehlt",
      "LEERER ABSCHNITT" in _text(build_pdf(_a, "Leer")), False)



leer = CrewPlan(date_from="", date_to="")
check("leere Planung ergibt trotzdem ein PDF", build_pdf(leer, "Leer")[:5], b"%PDF-")


# ─── 2. Excel ────────────────────────────────────────────────────────────────
print(chr(10) + "── 2. Excel ──")
from openpyxl import load_workbook                       # noqa: E402

x_kalk = build_xlsx(plan, "Schneider Electric @ HMI 2026", kalkulation=True)
x_kunde = build_xlsx(plan, "Schneider Electric @ HMI 2026", kalkulation=False)
ws_k = load_workbook(_io.BytesIO(x_kalk)).active
ws_c = load_workbook(_io.BytesIO(x_kunde)).active

# Die Kostenspalten stehen hinter den Tagen — bei 31 Tagen also ab Spalte 33.
_letzte = ws_k.max_column
spalten_k = [ws_k.cell(row=6, column=c).value for c in range(1, _letzte + 1)]
spalten_c = [ws_c.cell(row=6, column=c).value
             for c in range(1, ws_c.max_column + 1)]
check("Kalkulation hat Kostenspalten", "Summe" in spalten_k, True)
check("Kundenversion nicht", "Summe" in spalten_c, False)
check("Tagessatz nur in der Kalkulation",
      ("Tagessatz" in spalten_k, "Tagessatz" in spalten_c), (True, False))

# Der einzige Grund, Excel statt PDF zu geben, ist das Weiterrechnen — also gehören
# Formeln hinein und keine ausgerechneten Zahlen.
werte = [ws_k.cell(row=z, column=c).value
         for z in range(7, 12) for c in range(_letzte - 5, _letzte + 1)]
formeln = [w for w in werte if isinstance(w, str) and w.startswith("=")]
check("Kostenspalten als Formeln", len(formeln) > 0, True)
check("Summe rechnet aus der Zeile",
      any("*" in f and "+" in f for f in formeln), True)
check("Kopfzeilen und Namensspalte stehen fest", ws_k.freeze_panes is not None, True)

check("Excel ist eine ZIP-Datei", x_kalk[:2], b"PK")

# Dieselbe Regel wie im PDF: unbesetzte Zeilen kommen nicht in die Datei.
_wsx = load_workbook(_io.BytesIO(build_xlsx(_ohne, "Mit Leerzeile"))).active
_namen = [_wsx.cell(row=z, column=1).value for z in range(7, _wsx.max_row + 1)]
check("unbesetzte Zeile fehlt in der Excel", "Niemals dabei" in _namen, False)
check("besetzte Zeilen auch dort", "Ton-Operator" in _namen, True)


# ─── 3. Routen ───────────────────────────────────────────────────────────────
print(chr(10) + "── 3. Routen ──")
import db as _db                                          # noqa: E402

_db.init_db()
import routes.crew as crew                                # noqa: E402


class _S:
    """Sitzung mit dem Nötigsten — die Route liest nur Planung und Projektnamen."""
    def __init__(self):
        self.crew = plan
        self.crew_schedule = {}
        self.d83_project = None
        self.d83_name = "Schneider/Electric: HMI 2026.x83"


_ss = _S()


class _Req:
    session: dict = {}


crew.get_session = lambda _s: _ss          # type: ignore[assignment]
crew._lv_items = lambda _s: {}             # type: ignore[assignment]

# Dateiname: Projektnummer_Projektname_Personalplanung. Die Nummer vergibt Easyjob
# erst beim Anlegen — solange fällt sie weg. Die Kalkulation trägt zusätzlich ihren
# Namen: sie bleibt im Haus, und zwei gleich heißende Dateien im Downloadordner sind
# genau die Verwechslung, die man nicht bemerkt.
for art, kopf in (("pdf", b"%PDF-"), ("xlsx", b"PK")):
    for variante, teil in (("kalkulation", "Personalplanung_Kalkulation"),
                           ("kunde", "Personalplanung.")):
        r = run(crew.crew_export(_Req(), art=art, variante=variante))
        name = r.headers.get("content-disposition", "")
        check(f"{art}/{variante} liefert Daten", r.body[:len(kopf)], kopf)
        check(f"{art}/{variante} als Anhang", "attachment" in name, True)
        # Die Variante steht im Dateinamen: eine falsch verschickte Datei fällt
        # sonst niemandem auf.
        check(f"{art}/{variante} nach dem Muster benannt", teil in name, True)
        check(f"{art}/{variante} ohne Nummer im Import",
              "filename=\"Schneider" in name, True)
        # Aus dem Dateinamen müssen Schrägstriche und Doppelpunkte heraus, sonst
        # zerlegt der Browser den Pfad oder verweigert das Speichern.
        check(f"{art}/{variante} ohne Sonderzeichen",
              any(ch in name.split("filename=")[1] for ch in "/:\\"), False)

# Aus einem abgelegten Projekt geht der Export auch dann noch, wenn der Import
# längst vorbei ist — beim Nachreichen einer Beilage etwa.
with _db.get_conn() as _cn:
    _cn.execute("INSERT INTO projects (id, name, status, ej_project_number) "
                "VALUES (77, ?, 'uploaded', ?)",
                ("Porsche D Summit 2026", "26-0994"))
_db.save_crew_plan(77, plan.to_dict(), "test")
r = run(crew.crew_export(_Req(), art="pdf", variante="kunde", projekt_id=77))
check("Export aus dem Projekt", r.body[:5], b"%PDF-")
_nam = r.headers.get("content-disposition", "")
check("mit dem Projektnamen", "Porsche" in _nam, True)
# Beim abgelegten Projekt steht die Easyjob-Nummer vorn.
check("und der Projektnummer vorn", 'filename="26-0994_Porsche' in _nam, True)
check("und ohne Sitzung",
      len(r.body) > 1000, True)
r = run(crew.crew_export(_Req(), art="pdf", variante="kunde", projekt_id=999))
check("Projekt ohne Planung", r.status_code, 404)

r = run(crew.crew_export(_Req(), art="docx", variante="kalkulation"))
check("unbekanntes Format abgewiesen", r.status_code, 400)

_ss.crew = None
r = run(crew.crew_export(_Req(), art="pdf", variante="kalkulation"))
check("ohne Planung kein Export", r.status_code, 404)


print(chr(10) + "=" * 62)
if _fails:
    print(f"FEHLGESCHLAGEN: {len(_fails)}")
    for f in _fails:
        print("  -", f)
    sys.exit(1)
print("Alle Prüfungen bestanden.")
