"""Regressionstest der Preis-Rückschreibung in die Original-Excel.

Prüft, dass Preise in der richtigen Zelle landen, Gesamtpreis-Formeln unangetastet
bleiben und Formatierung/Struktur der Datei erhalten sind.

    .venv/Scripts/python.exe tests/test_excel_export.py
"""
import sys

# Konsole auf UTF-8: sonst stirbt schon ein "→" im print an cp1252 und der Test
# bricht mitten drin ab, ohne dass eine Prüfung fehlgeschlagen wäre.
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
os.chdir(os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

import excel_export as _xlout
import excel_parser as _xl

FN = "infos/PAG_NOC2026_LV Veranstaltungstechnik_V1.xlsx"
src = open(FN, "rb").read()
probe = _xl.probe_workbook(src)
layout = probe.layout
proj = _xl.parse_excel(src, layout)

sl = layout.sheet("01_Material")
print("Preisspalte 01_Material:", sl.roles.price, "(F=6 erwartet)")
assert sl.roles.price, "keine Preisspalte erkannt — Test sinnlos"

# Kunstpreise: jede Position aus 01_Material bekommt einen EP
mat = [i for i in proj.items if i.src_ref.startswith("01_Material!")]
prices = {i.item_id: round(100 + n * 1.5, 2) for n, i in enumerate(mat)}
print("Positionen 01_Material:", len(mat))

res = _xlout.write_prices(src, layout, proj, prices)
print(f"geschrieben={res.written} übersprungen={res.skipped} formeln={res.formulas}")
print("Hinweise:", res.notes or "keine")
assert res.written == len(mat), (res.written, len(mat))

# ── Gegenprobe: Datei wieder öffnen
wb_new = openpyxl.load_workbook(io.BytesIO(res.data), data_only=False)
wb_old = openpyxl.load_workbook(io.BytesIO(src), data_only=False)
ws_new, ws_old = wb_new["01_Material"], wb_old["01_Material"]

it16 = next(i for i in mat if i.src_ref == "01_Material!16")
want = prices[it16.item_id]
got  = ws_new.cell(16, sl.roles.price).value
print(f"Zeile 16: EP geschrieben = {got!r} (erwartet {want})")
assert got == want, (got, want)

# Gesamtpreis-Spalte G: im Original eine Formel? Dann unangetastet.
g_old, g_new = ws_old.cell(16, 7).value, ws_new.cell(16, 7).value
print(f"Zeile 16 Spalte G: vorher {g_old!r} -> nachher {g_new!r}")
if isinstance(g_old, str) and g_old.startswith("="):
    assert g_new == g_old, "Formel wurde überschrieben!"
    print("  → Formel belassen, Excel rechnet selbst ✓")
else:
    print("  → keine Formel, Gesamtpreis berechnet ✓")

# Struktur unverändert: alle Blätter da, Beschreibungen unangetastet
assert wb_new.sheetnames == wb_old.sheetnames, "Blätter verändert"
assert ws_new.cell(16, 2).value == ws_old.cell(16, 2).value, "Beschreibung verändert"
assert ws_new.cell(13, 2).value == ws_old.cell(13, 2).value, "Gruppenzeile verändert"
# Formatierung erhalten (Stichprobe: Fettschrift der Gruppenzeile)
assert bool(ws_new.cell(13, 2).font.bold) == bool(ws_old.cell(13, 2).font.bold), "Format verloren"
print("Blätter:", len(wb_new.sheetnames), "· Beschreibung/Gruppen/Format unverändert ✓")

# Zellen ohne Preis bleiben leer
untouched = ws_new.cell(13, sl.roles.price).value
print("Gruppenzeile 13 Preiszelle:", repr(untouched))
assert untouched == ws_old.cell(13, sl.roles.price).value

# ── Berechnete Preisspalte: Formeln dürfen NICHT überschrieben werden ───────
FN2 = "infos/LOS1_GOE Ausschreibung_Technik 2026_Preisblatt_V2.1.xlsx"
src2 = open(FN2, "rb").read()
pb2 = _xl.probe_workbook(src2)
lay2 = pb2.layout
sl2 = lay2.sheet("Preisblatt Technik")
for q in sl2.roles.qty:
    q.active = q.col in (8, 12)
proj2 = _xl.parse_excel(src2, lay2)

orig2 = openpyxl.load_workbook(io.BytesIO(src2), data_only=False)
formel = orig2["Preisblatt Technik"].cell(17, sl2.roles.price).value
print(f"\nGOE: erkannte Preisspalte {sl2.roles.price} enthält {formel!r}")
assert isinstance(formel, str) and formel.startswith("="), "Testannahme: E ist eine Formel"

p17 = [i for i in proj2.items if i.src_ref == "Preisblatt Technik!17"]
assert len(p17) == 2, len(p17)
res2 = _xlout.write_prices(src2, lay2, proj2, {p17[0].item_id: 50.0, p17[1].item_id: 80.0})
wb2 = openpyxl.load_workbook(io.BytesIO(res2.data), data_only=False)
nach = wb2["Preisblatt Technik"].cell(17, sl2.roles.price).value
print(f"  nach dem Export: {nach!r}")
assert nach == formel, f"Formel wurde überschrieben: {nach!r}"
assert res2.written == 0, f"nichts hätte geschrieben werden dürfen, sind {res2.written}"
assert any("Formeln" in n and "Spalte E" in n for n in res2.notes), res2.notes
print(f"  gemeldet: {[n for n in res2.notes if 'Formeln' in n][0][:96]}…")
print("  -> berechnete Spalte bleibt unangetastet und wird gemeldet ✓")

# ── Szenario-Kollision: zwei Szenarien, dieselbe Preiszelle ─────────────────
# PAG hat nur eine Mengenspalte — für den Test eine zweite aktivieren (Spalte H).
lay3 = _xl.probe_workbook(src).layout
sl3 = lay3.sheet("01_Material")
sl3.roles.qty = [_xl.QtyCol(col=5, label="A", job_name="A"),
                 _xl.QtyCol(col=8, label="B", job_name="B")]
proj3 = _xl.parse_excel(src, lay3)
z16 = [i for i in proj3.items if i.src_ref == "01_Material!16"]
print(f"\nPAG Zeile 16: {len(z16)} Szenario-Positionen -> {[i.item_id for i in z16]}")
assert len(z16) == 2, len(z16)
res3 = _xlout.write_prices(src, lay3, proj3, {z16[0].item_id: 50.0, z16[1].item_id: 80.0})
wb3 = openpyxl.load_workbook(io.BytesIO(res3.data))
val = wb3["01_Material"].cell(16, sl3.roles.price).value
print(f"  geschriebener EP = {val} (höherer von 50/80 erwartet)")
assert val == 80.0, val
assert any("mehrere Szenarien" in n for n in res3.notes), res3.notes
print(f"  Hinweis: {[n for n in res3.notes if 'mehrere' in n][0]}")

print(f"\nExcel-Rückschreibung ok.")
