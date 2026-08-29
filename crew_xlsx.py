"""Die Personalplanung als Excel — für Kollegen, die weiterrechnen wollen.

Dasselbe Bild wie im PDF, aber mit lebenden Zellen: die Summenspalten stehen als
**Formeln** in der Datei, nicht als ausgerechnete Zahlen. Wer eine Personenzahl oder
einen Satz ändert, sieht sofort, was das kostet — genau der Grund, aus dem die
Excel-Listen bisher überhaupt existierten.

``openpyxl`` ist bereits Abhängigkeit (Excel-Import des LV), kommt hier also ohne
neues Paket aus.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from crew_plan import CrewPlan, parse_day

GELB = "FFDC00"
KOPF_GRAU = "ECEFF1"
BAND_GRAU = "EEF1F6"
WE_GRAU = "F4F6F8"
SUMME_GRAU = "F7F8FC"

_WOCHENTAG = ("Mo", "Di", "Mi", "Do", "Fr", "Sa", "So")
_MONAT = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli",
          "August", "September", "Oktober", "November", "Dezember")

PHASENFARBEN = ["1565C0", "00897B", "E65100", "6A1B9A",
                "2E7D32", "C62828", "455A64"]

_DUENN = Side(style="thin", color="D8DEE4")
_RAHMEN = Border(left=_DUENN, right=_DUENN, top=_DUENN, bottom=_DUENN)
_EURO = '#.##0 "€"'


def _gefuellte_gruppen(plan: CrewPlan):
    """Nur Abschnitte mit besetzten Zeilen — wie im PDF.

    Zeilen ohne Manntage stehen in der Matrix als Vorschlag aus dem Matching; in der
    ausgegebenen Liste wären sie nur Leute, die nicht kommen.
    """
    out = []
    for key, rows in plan.groups():
        gefuellt = [r for r in rows if plan.manntage(r) > 0]
        if gefuellt:
            out.append((key, gefuellt))
    return out


def _phase_von(plan: CrewPlan, key: str) -> int:
    for i, ph in enumerate(plan.phases):
        if ph.contains(key):
            return i
    return -1


def build_xlsx(plan: CrewPlan, projekt: str = "", *, kalkulation: bool = True,
               menu_titel: dict[str, str] | None = None) -> bytes:
    menu_titel = menu_titel or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Personalplanung"

    tage = plan.day_keys()
    n_tage = len(tage)
    spalte_tag1 = 2                      # A = Name, ab B die Tage
    spalte_nach = spalte_tag1 + n_tage   # erste Kostenspalte

    kosten = (["MT", "Tagessatz", "Spesen", "Hotel", "Reisen", "Summe"]
              if kalkulation else [])

    # ── Titel ────────────────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value="Personalplanung").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=projekt or "Ohne Projektnamen").font = Font(size=11)
    if plan.date_from and plan.date_to:
        a, b = parse_day(plan.date_from), parse_day(plan.date_to)
        ws.cell(row=2, column=spalte_tag1,
                value=(f"{a.day}. {_MONAT[a.month - 1]} – "
                       f"{b.day}. {_MONAT[b.month - 1]} {b.year}"))

    # ── Kopfzeilen: Phase, Wochentag, Tag ────────────────────────────────────
    z_phase, z_wt, z_tag = 4, 5, 6

    lauf = 0
    while lauf < n_tage:
        pi = _phase_von(plan, tage[lauf])
        span = 0
        while lauf + span < n_tage and _phase_von(plan, tage[lauf + span]) == pi:
            span += 1
        if pi >= 0:
            c1 = spalte_tag1 + lauf
            zelle = ws.cell(row=z_phase, column=c1, value=plan.phases[pi].name.upper())
            zelle.font = Font(bold=True, size=8, color="FFFFFF")
            zelle.alignment = Alignment(horizontal="center")
            farbe = PHASENFARBEN[pi % len(PHASENFARBEN)]
            for i in range(span):
                ws.cell(row=z_phase, column=c1 + i).fill = PatternFill(
                    "solid", fgColor=farbe)
            if span > 1:
                ws.merge_cells(start_row=z_phase, start_column=c1,
                               end_row=z_phase, end_column=c1 + span - 1)
        lauf += span

    ws.cell(row=z_tag, column=1, value="Ressource").font = Font(bold=True, size=9)
    for i, key in enumerate(tage):
        d = parse_day(key)
        we = d.weekday() >= 5
        for zeile, wert, fett in ((z_wt, _WOCHENTAG[d.weekday()], False),
                                  (z_tag, d.day, True)):
            z = ws.cell(row=zeile, column=spalte_tag1 + i, value=wert)
            z.font = Font(bold=fett, size=8, color="90A4AE" if we else "37474F")
            z.alignment = Alignment(horizontal="center")
            z.fill = PatternFill("solid", fgColor=WE_GRAU if we else KOPF_GRAU)

    for j, name in enumerate(kosten):
        z = ws.cell(row=z_tag, column=spalte_nach + j, value=name)
        z.font = Font(bold=True, size=9)
        z.alignment = Alignment(horizontal="right")
        z.fill = PatternFill("solid", fgColor=KOPF_GRAU)

    # ── Zeilen ───────────────────────────────────────────────────────────────
    zeile = z_tag + 1
    daten_zeilen: list[int] = []
    for menu_key, rows in _gefuellte_gruppen(plan):
        titel = (menu_titel.get(menu_key) or plan.menu_title(menu_key)
                 or "Ohne Abschnitt")
        z = ws.cell(row=zeile, column=1, value=titel.upper())
        z.font = Font(bold=True, size=9, color="37474F")
        for sp in range(1, spalte_nach + len(kosten)):
            ws.cell(row=zeile, column=sp).fill = PatternFill("solid",
                                                             fgColor=BAND_GRAU)
        zeile += 1

        for r in rows:
            daten_zeilen.append(zeile)
            ws.cell(row=zeile, column=1, value=r.label).font = Font(size=9)
            for i, key in enumerate(tage):
                d = parse_day(key)
                n = r.cells.get(key)
                z = ws.cell(row=zeile, column=spalte_tag1 + i, value=n or None)
                z.alignment = Alignment(horizontal="center")
                z.font = Font(bold=True, size=9)
                z.border = _RAHMEN
                if n:
                    z.fill = PatternFill("solid", fgColor=GELB)
                elif d.weekday() >= 5:
                    z.fill = PatternFill("solid", fgColor=WE_GRAU)

            if kalkulation:
                _kostenzeile(ws, plan, r, zeile, spalte_tag1, spalte_nach, n_tage)
            zeile += 1

    # ── Summen ───────────────────────────────────────────────────────────────
    z_summe = zeile
    ws.cell(row=z_summe, column=1, value="Personen / Tag").font = Font(bold=True, size=9)
    for i in range(n_tage):
        sp = get_column_letter(spalte_tag1 + i)
        z = ws.cell(row=z_summe, column=spalte_tag1 + i)
        if daten_zeilen:
            z.value = (f"=SUM({sp}{daten_zeilen[0]}:{sp}{daten_zeilen[-1]})")
        z.font = Font(bold=True, size=9)
        z.alignment = Alignment(horizontal="center")
    for sp in range(1, spalte_nach + len(kosten)):
        ws.cell(row=z_summe, column=sp).fill = PatternFill("solid", fgColor=SUMME_GRAU)

    if kalkulation and daten_zeilen:
        for j, name in enumerate(kosten):
            if name == "Tagessatz":
                continue
            sp = get_column_letter(spalte_nach + j)
            z = ws.cell(row=z_summe, column=spalte_nach + j,
                        value=f"=SUM({sp}{daten_zeilen[0]}:{sp}{daten_zeilen[-1]})")
            z.font = Font(bold=True, size=9)
            if name != "MT":
                z.number_format = _EURO

    # ── Maße ─────────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    for i in range(n_tage):
        ws.column_dimensions[get_column_letter(spalte_tag1 + i)].width = 4.2
    for j, name in enumerate(kosten):
        ws.column_dimensions[get_column_letter(spalte_nach + j)].width = (
            6 if name == "MT" else 11)
    # Namensspalte und Kopfzeilen stehen fest: ohne das verliert man beim Scrollen
    # nach rechts die Zeilenbeschriftung, und genau dafür ist die Liste da.
    ws.freeze_panes = ws.cell(row=z_tag + 1, column=spalte_tag1)

    puffer = io.BytesIO()
    wb.save(puffer)
    return puffer.getvalue()


def _kostenzeile(ws, plan: CrewPlan, r, zeile: int, spalte_tag1: int,
                 spalte_nach: int, n_tage: int) -> None:
    """Die Kostenspalten als Formeln.

    Ausgerechnete Zahlen wären hier verschenkt: der einzige Grund, die Liste als Excel
    statt als PDF zu geben, ist das Weiterrechnen. Die Spesenformel bildet die Regel
    ab — voller Satz für Tage mit Übernachtung, halber für Tage ohne.
    """
    von = get_column_letter(spalte_tag1)
    bis = get_column_letter(spalte_tag1 + n_tage - 1)
    mt = get_column_letter(spalte_nach)
    ts = get_column_letter(spalte_nach + 1)
    spesen = get_column_letter(spalte_nach + 2)
    hotel = get_column_letter(spalte_nach + 3)
    reisen = get_column_letter(spalte_nach + 4)

    naechte = plan.naechte(r)
    satz = plan.spesen_satz
    werte = [
        (spalte_nach, f"=SUM({von}{zeile}:{bis}{zeile})", None),
        (spalte_nach + 1, r.tagessatz, _EURO),
        # Nächte stecken in der Formel, damit sie sichtbar bleiben: MT − N halbe
        # Sätze plus N volle.
        (spalte_nach + 2,
         f"=({mt}{zeile}-{naechte})*{satz / 2:g}+{naechte}*{satz:g}", _EURO),
        (spalte_nach + 3, f"={naechte}*{plan.hotel_satz_of(r):g}", _EURO),
        (spalte_nach + 4, f"={r.rk_anzahl}*{plan.rk_satz_of(r):g}", _EURO),
        (spalte_nach + 5,
         f"={mt}{zeile}*{ts}{zeile}+{spesen}{zeile}+{hotel}{zeile}+{reisen}{zeile}",
         _EURO),
    ]
    for sp, wert, format_ in werte:
        z = ws.cell(row=zeile, column=sp, value=wert)
        z.font = Font(size=9, bold=sp == spalte_nach + 5)
        if format_:
            z.number_format = format_
